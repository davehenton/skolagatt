from django.shortcuts import render, redirect
from django.views.generic import (
    ListView, CreateView, DetailView, UpdateView, DeleteView
)
from django.contrib.auth.mixins import UserPassesTestMixin

from django.core.urlresolvers import reverse_lazy
import openpyxl
from ast import literal_eval

from . import forms
from .models import (
    Survey, SurveyType, SurveyResource, SurveyTransformation,
    SurveyGradingTemplate, SurveyInputField, SurveyInputGroup
)
import common.util as common_util
from common.models import (
    SurveyResult,
    StudentGroup,
    GroupSurvey,
)
from .mixins import (
    SurveySuperSuccessMixin,
    SurveyCreateSuperSuccessMixin,
    SurveyDeleteSuperSuccessMixin
)


class SurveyListing(UserPassesTestMixin, ListView):
    model = Survey

    def test_func(self):
        return self.request.user.is_authenticated()


class SurveyDetail(UserPassesTestMixin, DetailView):
    model = Survey

    def get_context_data(self, **kwargs):
        context = super(SurveyDetail, self).get_context_data(**kwargs)
        survey = Survey.objects.get(pk=self.kwargs['pk'])
        context['survey'] = survey
        context['survey_resource_list'] = SurveyResource.objects.filter(survey=survey)
        context['survey_template_list'] = SurveyGradingTemplate.objects.filter(survey=survey)
        context['survey_transformation_list'] = SurveyTransformation.objects.filter(survey=survey)
        input_groups = SurveyInputGroup.objects.filter(survey=survey)
        context['survey_input_field_groups'] = input_groups
        inputs = []
        for group in input_groups:
            inputs.extend(SurveyInputField.objects.filter(
                input_group=group
            ))
        context['survey_input_field_list'] = inputs
        return context

    def test_func(self, **kwargs):
        return self.request.user.is_authenticated()


class SurveyTypeDetail(UserPassesTestMixin, DetailView):
    model = SurveyType

    def get_context_data(self, **kwargs):
        context = super(SurveyTypeDetail, self).get_context_data(**kwargs)
        context['survey_type'] = SurveyType.objects.get(pk=self.kwargs['pk'])
        return context

    def test_func(self, **kwargs):
        return self.request.user.is_authenticated()


class SurveyTypeCreate(SurveyCreateSuperSuccessMixin, CreateView):
    model = SurveyType
    form_class = forms.SurveyTypeForm

    def form_valid(self, form):
        survey = form.save(commit=False)
        survey.created_by = self.request.user
        return super(SurveyTypeCreate, self).form_valid(form)

    def get_success_url(self):
        return reverse_lazy('survey:survey_list')


class SurveyTypeUpdate(SurveySuperSuccessMixin, UpdateView):
    model = SurveyType
    form_class = forms.SurveyTypeForm

    def get_success_url(self):
        return reverse_lazy('survey:survey_list')


class SurveyTypeDelete(UserPassesTestMixin, DeleteView):
    model = SurveyType
    success_url = reverse_lazy('survey:survey_list')
    login_url = reverse_lazy('denied')
    template_name = "survey/confirm_delete.html"

    def test_func(self):
        return self.request.user.is_superuser


class SurveyCreate(SurveyCreateSuperSuccessMixin, CreateView):
    model = Survey
    form_class = forms.SurveyForm

    def form_valid(self, form):
        survey = form.save(commit=False)
        survey.created_by = self.request.user
        return super(SurveyCreate, self).form_valid(form)


class SurveyUpdate(SurveySuperSuccessMixin, UpdateView):
    model = Survey
    form_class = forms.SurveyForm


class SurveyDelete(UserPassesTestMixin, DeleteView):
    model = Survey
    success_url = reverse_lazy('survey:survey_list')
    login_url = reverse_lazy('denied')
    template_name = "survey/confirm_delete.html"

    def test_func(self):
        return self.request.user.is_superuser


class SurveyResourceDetail(UserPassesTestMixin, DetailView):
    model = SurveyResource

    def test_func(self, **kwargs):
        return True

    def get_context_data(self, **kwargs):
        # xxx will be available in the template as the related objects
        context = super(SurveyResourceDetail, self).get_context_data(**kwargs)
        context['survey'] = Survey.objects.get(pk=self.kwargs['survey_id'])
        return context


class SurveyResourceCreate(SurveySuperSuccessMixin, CreateView):
    model = SurveyResource
    form_class = forms.SurveyResourceForm

    def form_valid(self, form):
        survey = form.save(commit=False)
        survey.created_by = self.request.user
        survey.survey = Survey.objects.get(pk=self.kwargs['survey_id'])
        return super(SurveyResourceCreate, self).form_valid(form)

    def get_context_data(self, **kwargs):
        context = super(SurveyResourceCreate, self).get_context_data(**kwargs)
        context['survey'] = Survey.objects.get(pk=self.kwargs['survey_id'])
        return context


class SurveyResourceUpdate(SurveySuperSuccessMixin, UpdateView):
    model = SurveyResource
    form_class = forms.SurveyResourceForm


class SurveyResourceDelete(SurveyDeleteSuperSuccessMixin, DeleteView):
    model = SurveyResource


class SurveyGradingTemplateDetail(UserPassesTestMixin, DetailView):
    model = SurveyGradingTemplate

    def test_func(self, **kwargs):
        return True

    def get_context_data(self, **kwargs):
        # xxx will be available in the template as the related objects
        context = super(SurveyGradingTemplateDetail, self).get_context_data(**kwargs)
        context['survey'] = Survey.objects.get(pk=self.kwargs['survey_id'])
        return context


class SurveyGradingTemplateCreate(SurveySuperSuccessMixin, CreateView):
    model = SurveyGradingTemplate
    form_class = forms.SurveyGradingTemplateForm

    def form_valid(self, form):
        survey = form.save(commit=False)
        survey.created_by = self.request.user
        survey.survey = Survey.objects.get(pk=self.kwargs['survey_id'])
        return super(SurveyGradingTemplateCreate, self).form_valid(form)

    def get_context_data(self, **kwargs):
        context = super(SurveyGradingTemplateCreate, self).get_context_data(**kwargs)
        context['survey'] = Survey.objects.get(pk=self.kwargs['survey_id'])
        return context


class SurveyGradingTemplateUpdate(SurveySuperSuccessMixin, UpdateView):
    model = SurveyGradingTemplate
    form_class = forms.SurveyGradingTemplateForm


class SurveyGradingTemplateDelete(SurveyDeleteSuperSuccessMixin, DeleteView):
    model = SurveyGradingTemplate


class SurveyInputFieldDetail(UserPassesTestMixin, DetailView):
    model = SurveyInputField

    def test_func(self, **kwargs):
        return True

    def get_context_data(self, **kwargs):
        # xxx will be available in the template as the related objects
        context = super(SurveyInputFieldDetail, self).get_context_data(**kwargs)
        context['survey'] = Survey.objects.get(pk=self.kwargs['survey_id'])
        return context


class SurveyInputFieldCreate(SurveySuperSuccessMixin, CreateView):
    model = SurveyInputField
    form_class = forms.SurveyInputFieldForm

    def form_valid(self, form):
        survey = form.save(commit=False)
        survey.created_by = self.request.user
        survey.survey = Survey.objects.get(pk=self.kwargs['survey_id'])
        return super(SurveyInputFieldCreate, self).form_valid(form)

    def get_context_data(self, **kwargs):
        context = super(SurveyInputFieldCreate, self).get_context_data(**kwargs)
        survey = Survey.objects.get(pk=self.kwargs['survey_id'])
        context['survey'] = survey
        # Only get input groups related to the survey object for the select menu
        context['form'].fields['input_group'].queryset = SurveyInputGroup.objects.filter(
            survey=survey
        )
        return context


class SurveyInputFieldUpdate(SurveySuperSuccessMixin, UpdateView):
    model = SurveyInputField
    form_class = forms.SurveyInputFieldForm

    def get_context_data(self, **kwargs):
        context = super(SurveyInputFieldUpdate, self).get_context_data(**kwargs)
        survey = Survey.objects.get(pk=self.kwargs['survey_id'])
        context['survey'] = survey
        # Only get input groups related to the survey object for the select menu
        context['form'].fields['input_group'].queryset = SurveyInputGroup.objects.filter(
            survey=survey
        )
        return context


class SurveyInputFieldDelete(SurveyDeleteSuperSuccessMixin, DeleteView):
    model = SurveyInputField


class SurveyInputGroupDetail(UserPassesTestMixin, DetailView):
    model = SurveyInputGroup

    def test_func(self, **kwargs):
        return True

    def get_context_data(self, **kwargs):
        # xxx will be available in the template as the related objects
        context = super(SurveyInputGroupDetail, self).get_context_data(**kwargs)
        context['survey'] = Survey.objects.get(pk=self.kwargs['survey_id'])
        return context


class SurveyInputGroupCreate(SurveySuperSuccessMixin, CreateView):
    model = SurveyInputGroup
    form_class = forms.SurveyInputGroupCreateForm

    def form_valid(self, form):
        input_group = form.save(commit=False)
        input_group.survey = Survey.objects.get(pk=self.kwargs['survey_id'])
        input_group.save()  # Save input group now so we can link inputfields to it
        # Create related SurveyInputFields
        nr_inputs = int(self.request.POST['inputs'])
        for x in range(0, nr_inputs):
            inputfield = SurveyInputField(
                name=x + 1, label=input_group.title, input_group=input_group
            )
            inputfield.save()

        return super(SurveyInputGroupCreate, self).form_valid(form)

    def get_context_data(self, **kwargs):
        context = super(SurveyInputGroupCreate, self).get_context_data(**kwargs)
        context['survey'] = Survey.objects.get(pk=self.kwargs['survey_id'])
        return context


class SurveyInputGroupUpdate(SurveySuperSuccessMixin, UpdateView):
    model = SurveyInputGroup
    form_class = forms.SurveyInputGroupForm

    def form_valid(self, form):
        input_group = form.save(commit=False)
        input_group.survey = Survey.objects.get(pk=self.kwargs['survey_id'])
        input_group.save()  # Save input group now so we can link inputfields to it
        # Create related SurveyInputFields
        nr_inputs = int(self.request.POST['inputs'])
        nr_inputs_saved = input_group.num_input_fields()
        if nr_inputs_saved > nr_inputs:
            for x in range(nr_inputs, nr_inputs_saved):
                SurveyInputField.objects.get(input_group=input_group, name=x + 1).delete()
        elif nr_inputs_saved < nr_inputs:
            for x in range(0, nr_inputs):
                SurveyInputField.objects.get_or_create(
                    name=x + 1, label=input_group.title, input_group=input_group
                )

        return super(SurveyInputGroupUpdate, self).form_valid(form)


class SurveyInputGroupDelete(SurveyDeleteSuperSuccessMixin, DeleteView):
    model = SurveyInputGroup


# Transformation


class SurveyTransformationDetail(UserPassesTestMixin, DetailView):
    model = SurveyTransformation

    def get_context_data(self, **kwargs):
        # xxx will be available in the template as the related objects
        context = super(SurveyTransformationDetail, self).get_context_data(**kwargs)
        context['survey'] = Survey.objects.get(pk=self.kwargs['survey_id'])
        transformation = SurveyTransformation.objects.get(pk=self.kwargs['pk'])
        context['transformation'] = transformation
        return context

    def test_func(self, **kwargs):
        return self.request.user.is_authenticated()


class SurveyTransformationCreate(SurveySuperSuccessMixin, CreateView):
    model = SurveyTransformation
    form_class = forms.SurveyTransformationForm

    def get_context_data(self, **kwargs):
        context = super(SurveyTransformationCreate, self).get_context_data(**kwargs)
        context['survey'] = Survey.objects.get(pk=self.kwargs['survey_id'])
        return context

    def post(self, *args, **kwargs):
        if(self.request.FILES):
            try:
                input_excel = self.request.FILES['file']
                book = openpyxl.load_workbook(filename=input_excel)
                name = self.request.POST.get('transformationname')
                unit = self.request.POST.get('transformationunit')
                survey = self.kwargs['survey_id']
                result_data = {}
                for sheet in book.worksheets:
                    # import pdb; pdb.set_trace()
                    for index, row in enumerate(
                            sheet.iter_rows('A{}:B{}'.format(sheet.min_row + 1, sheet.max_row))):
                        # result_data[row] = float(sheet.cell(row, 1))
                        result_data[int(sheet.cell(row=index + 2, column=1).value)] = float(
                            sheet.cell(row=index + 2, column=2).value)
                # order select
                transform = SurveyTransformation(name=name, order=1, data=result_data, unit=unit)
                transform.survey = Survey.objects.get(id=survey)
                transform.save()

            except Exception as e:
                return render(
                    self.request,
                    'survey/surveytransformation_form.html',
                    {'pk': self.kwargs['survey_id']},
                    {'error': 'Dálkur ekki til, reyndu aftur'}
                )
        return redirect(self.get_success_url())

    def get_success_url(self):
        if 'school_id' in self.kwargs:
            return reverse_lazy(
                'samraemd:math_listing',
                kwargs={'pk': self.kwargs['survey_id']}
            )
        return reverse_lazy('survey:survey_detail',
                            kwargs={'pk': self.kwargs['survey_id']})


class SurveyTransformationDelete(SurveyDeleteSuperSuccessMixin, DeleteView):
    model = SurveyTransformation
