# -*- coding: utf-8 -*-
from django.http import HttpResponse

from ast import literal_eval

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.chart import AreaChart, BarChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout

from survey.models import (
    Survey,
)
from common.models import (
    School,
    StudentGroup,
    GroupSurvey,
    SurveyResult,
)


def lesfimi_excel_entire_country_stats():
    ref_values = {
        1: (20, 55, 75),
        2: (40, 85, 100),
        3: (55, 100, 120),
        4: (80, 120, 145),
        5: (90, 140, 160),
        6: (105, 155, 175),
        7: (120, 165, 190),
        8: (130, 180, 210),
        9: (140, 180, 210),
        10: (145, 180, 210),
    }

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Lesfimi - Allt Landið.xlsx'

    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    wb.remove_sheet(ws)

    tests = (
        ('b{}_LF_mai17', 'Maí 2017'),
        ('b{}_LF_jan17', 'Janúar 2017'),
        ('{}b_LF_sept', 'September 2016'),
    )
    for test in tests:
        identifier = test[0]
        title = test[1]
        ws = wb.create_sheet(title=title)
        ws['A1'] = 'Árgangur'
        ws['B1'] = 'Fjöldi nemenda'
        ws['C1'] = 'Fjöldi sem þreytti próf'
        ws['D1'] = 'Hlutfall sem þreytti próf'
        ws['E1'] = 'Hlutfall sem nær 90% viðmiðum'
        ws['F1'] = 'Hlutfall sem nær 50% viðmiðum'
        ws['G1'] = 'Hlutfall sem nær 25% viðmiðum'
        index = 2
        errors = []
        for year in range(1, 11):
            ws['A' + str(index)] = year
            survey = Survey.objects.filter(identifier=identifier.format(year)).first()
            studentgroups = StudentGroup.objects.filter(student_year=year).all()
            this_year_result = {
                'students': 0,
                'students_who_took_test': 0,
                'students_over_25pct': 0,
                'students_over_50pct': 0,
                'students_over_90pct': 0,
            }
            for studentgroup in studentgroups:
                this_year_result['students'] += studentgroup.students.all().count()
                groupsurveys = GroupSurvey.objects.filter(studentgroup=studentgroup, survey=survey)
                if groupsurveys.all().count() > 1:
                    errors.append('sama próf skráð {} sinnum fyrir {} í {}'.format(
                        groupsurveys.all().count(), studentgroup.name, studentgroup.school.name))
                for groupsurvey in groupsurveys.all():
                    for student in studentgroup.students.all():
                        surveyresults = SurveyResult.objects.filter(survey=groupsurvey, student=student)
                        if surveyresults.all().count() > 1:
                            errors.append('{} niðurstöður í sama prófi skráðar fyrir nemanda {} í bekk {} í {}'.format(
                                surveyresults.all().count(),
                                student.ssn,
                                studentgroup.name,
                                studentgroup.school.name,
                            ))
                        for surveyresult in surveyresults.all():
                            try:
                                survey_student_result = surveyresult.calculated_results()
                                if not survey_student_result[0] == '':
                                    this_year_result['students_who_took_test'] += 1
                                    if int(survey_student_result[0]) >= ref_values[year][2]:
                                        this_year_result['students_over_25pct'] += 1
                                    if int(survey_student_result[0]) >= ref_values[year][1]:
                                        this_year_result['students_over_50pct'] += 1
                                    if int(survey_student_result[0]) >= ref_values[year][0]:
                                        this_year_result['students_over_90pct'] += 1
                            except:
                                pass

            ws['B' + str(index)] = this_year_result['students']
            ws['C' + str(index)] = this_year_result['students_who_took_test']
            if this_year_result['students'] > 0 and this_year_result['students_who_took_test'] > 0:
                ws['D' + str(index)] = (this_year_result['students_who_took_test'] / this_year_result['students']) * 100
                pct_over_90pct = (this_year_result['students_over_90pct'] /
                                  this_year_result['students_who_took_test']) * 100
                ws['E' + str(index)] = pct_over_90pct

                pct_over_50pct = (this_year_result['students_over_50pct'] /
                                  this_year_result['students_who_took_test']) * 100
                ws['F' + str(index)] = pct_over_50pct

                pct_over_25pct = (this_year_result['students_over_25pct'] /
                                  this_year_result['students_who_took_test']) * 100
                ws['G' + str(index)] = pct_over_25pct

            else:
                ws['D' + str(index)] = 0
                ws['E' + str(index)] = 0
                ws['F' + str(index)] = 0
                ws['G' + str(index)] = 0

            index += 1
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
        for col, value in dims.items():
            ws.column_dimensions[col].width = int(value) + 2

        chart = AreaChart()
        chart.title = "Lesfimi í {} - Allt landið".format(title)
        chart.style = 10
        chart.width = 40
        chart.height = 20

        chart.layout = Layout(
            ManualLayout(
                xMode="edge",
                yMode="edge",
            )
        )

        chart.x_axis.title = 'Árgangur'
        chart.y_axis.title = 'Prósent'

        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 100

        cats = Reference(ws, min_col=1, min_row=2, max_row=index - 1)
        data = Reference(ws, min_col=5, min_row=1, max_col=7, max_row=index)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        bchart = BarChart()
        bchart.title = "Hlutfall nemenda sem þreyttu próf"
        bchart.style = 10
        bchart.width = 20
        bchart.height = 10

        bchart.x_axis.title = 'Árgangur'
        bchart.y_axis.title = 'Prósent'

        bchart.y_axis.scaling.min = 0
        bchart.y_axis.scaling.max = 100

        bdata = Reference(ws, min_col=4, max_col=4, min_row=2, max_row=index)
        bchart.add_data(bdata)
        bchart.legend = None
        bchart.set_categories(cats)
        ws.add_chart(bchart, "I1")

        if index > 20:
            ws.add_chart(chart, "A" + str(index + 2))
        else:
            ws.add_chart(chart, "A22")

        if errors:
            ws = wb.create_sheet(title='Villur')
            index = 1
            for error in errors:
                ws['A' + str(index)] = "ATH: " + error
                ws['A' + str(index)].fill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')
                ws.merge_cells('A' + str(index) + ':F' + str(index))
                index += 1

    wb.save(filename='/tmp/test.xlsx')
    # wb.save(response)

    return response


def _samraemd_excel_maeting():
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()

    schools = School.objects.all()
    surveys = Survey.objects.filter(
        identifier__in=['SKP109', 'SKP110', 'SKP209', 'SKP210'],
    )
    ws['A1'] = 'Kennitala'
    ws['B1'] = 'Skóli'
    ws['C1'] = 'Skólanúmer'
    ws['D1'] = 'Próf'
    ws['E1'] = 'Kóði'
    ws['F1'] = 'Staða'

    index = 2
    for school in schools:
        studentgroups = StudentGroup.objects.filter(
            school=school,
            student_year__in=['9', '10'],
        ).all()
        for studentgroup in studentgroups:
            groupsurveys = GroupSurvey.objects.filter(
                studentgroup=studentgroup,
                survey__in=surveys,
            ).all()
            for groupsurvey in groupsurveys:
                survey_results = SurveyResult.objects.filter(
                    survey=groupsurvey,
                ).all()
                for sr in survey_results:
                    r = literal_eval(sr.results)
                    ws['A' + str(index)] = sr.student.ssn
                    ws['B' + str(index)] = studentgroup.school.name
                    ws['C' + str(index)] = studentgroup.school.school_nr
                    ws['D' + str(index)] = groupsurvey.survey.identifier
                    try:
                        click_values = literal_eval(r['click_values'])
                        vals = click_values[0].split(',')
                        ws['E' + str(index)] = vals[0]
                        ws['F' + str(index)] = vals[1]
                    except:
                        ws['E' + str(index)] = ''
                    index += 1
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = int(value) + 2
    wb.save(filename='/tmp/maeting.xlsx')


def _lesfimi_excel_result_duplicates():
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    wb.remove_sheet(ws)

    tests = (
        ('b{}_LF_jan17', 'Janúar 2017'),
        ('{}b_LF_sept', 'September 2016'),
    )
    for test in tests:
        identifier = test[0]
        title = test[1]
        ws = wb.create_sheet(title=title)
        ws['A1'] = 'Kennitala'
        ws['B1'] = 'Groupsurvey id'
        ws['C1'] = 'Surveyresult id'
        ws['D1'] = 'Result'
        index = 2
        errors = []
        for year in range(1, 11):
            ws['A' + str(index)] = year
            survey = Survey.objects.filter(identifier=identifier.format(year)).first()
            survey_type = SurveyType.objects.filter(survey=survey.id).values('id')
            dic = survey_type[0]
            survey_type = dic['id']
            studentgroups = StudentGroup.objects.filter(student_year=year).all()
            for studentgroup in studentgroups:
                groupsurveys = GroupSurvey.objects.filter(studentgroup=studentgroup, survey=survey)
                if groupsurveys.all().count() > 1:
                    errors.append('sama próf skráð {} sinnum fyrir {} í {}'.format(
                        groupsurveys.all().count(), studentgroup.name, studentgroup.school.name))
                for groupsurvey in groupsurveys.all():
                    for student in studentgroup.students.all():
                        surveyresults = SurveyResult.objects.filter(survey=groupsurvey, student=student)
                        if surveyresults.all().count() > 1:
                            for surveyresult in surveyresults.all():
                                ws['A' + str(index)] = student.ssn
                                ws['B' + str(index)] = groupsurvey.id
                                ws['C' + str(index)] = surveyresult.id
                                ws['D' + str(index)] = surveyresult.results
                                index += 1

    wb.save(filename='/tmp/duplicates.xlsx')


def _surveyresult_dupls():
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws['A1'] = 'surveyresult_id'
    ws['B1'] = 'student_id'
    ws['C1'] = 'survey_id'
    ws['D1'] = 'results'
    groupsurveys = GroupSurvey.objects.all()
    index = 2
    for groupsurvey in groupsurveys:
        if groupsurvey.studentgroup:
            for student in groupsurvey.studentgroup.students.all():
                results = SurveyResult.objects.filter(survey=groupsurvey, student=student)
                if (results.count() > 1):
                    for result in results.all():
                        ws['A' + str(index)] = str(result.id)
                        ws['B' + str(index)] = str(result.student_id)
                        ws['C' + str(index)] = str(result.survey_id)
                        ws['D' + str(index)] = str(result.results)
                        index += 1
                    index += 1
    wb.save(filename='/tmp/dupl.xlsx')


def _generate_excel_audun():
    tests = [
        ['{}b_LF_sept', 'September 2016', None, None],
        ['b{}_LF_jan17', 'Janúar 2017', None, None],
        ['b{}_LF_mai17', 'Maí 2017', None, None],
    ]

    wb = openpyxl.Workbook()

    ws = wb.get_active_sheet()
    wb.remove_sheet(ws)

    for year in range(1, 11):
        # Get all the surveys
        tests_year = tests
        for i in range(0, len(tests_year)):
            identifier = tests_year[i][0].format(year)
            survey = Survey.objects.get(identifier=identifier)
            tests_year[i][2] = survey

        title = "Árgangur {}".format(year)
        ws = wb.create_sheet(title=title)
        ws['A1'] = 'Kennitala nemanda'
        ws['B1'] = 'Nafn'
        ws['C1'] = 'Nafn skóla'
        ws['D1'] = 'Skóla nr'
        ws['E1'] = 'September'
        ws['F1'] = 'Janúar'
        ws['G1'] = 'Maí'
        ws['H1'] = 'Mismunur'
        ws['I1'] = 'September óþjálgað'
        ws['J1'] = 'Janúar óþjálgað'
        ws['K1'] = 'Maí óþjálgað'
        index = 2
        groups = StudentGroup.objects.filter(student_year=year).all()
        for group in groups:
            school = group.school
            tests_group = tests_year
            for i in range(0, len(tests_group)):
                survey = tests_group[i][2]
                try:
                    groupsurvey = GroupSurvey.objects.get(studentgroup=group, survey=survey)
                except GroupSurvey.MultipleObjectsReturned:
                    groupsurvey = GroupSurvey.objects.filter(studentgroup=group, survey=survey).first()
                except GroupSurvey.DoesNotExist:
                    groupsurvey = None
                tests_group[i][3] = groupsurvey

            for student in group.students.all():
                ws['A' + str(index)] = student.ssn
                ws['B' + str(index)] = student.name
                ws['C' + str(index)] = school.name
                ws['D' + str(index)] = school.school_nr
                col = ord('E')
                col_nt = ord('I')
                for test in tests_group:
                    groupsurvey = test[3]

                    if groupsurvey:
                        try:
                            result = SurveyResult.objects.get(student=student, survey=groupsurvey)
                            calc_res = result.calculated_results()[0]
                            calc_res_nt = result.calculated_results(use_transformation=False)[0]
                        except SurveyResult.DoesNotExist:
                            calc_res = 'N/A'
                            calc_res_nt = 'N/A'
                        except SurveyResult.MultipleObjectsReturned:
                            result = SurveyResult.objects.filter(
                                student=student,
                                survey=groupsurvey
                            ).order_by('-id')
                            calc_res = result.first().calculated_results()[0]
                            calc_res_nt = result.first().calculated_results(use_transformation=False)[0]
                    else:
                        calc_res = 'N/A'
                        calc_res_nt = 'N/A'

                    ws[chr(col) + str(index)] = calc_res
                    ws[chr(col_nt) + str(index)] = calc_res_nt
                    col += 1
                    col_nt += 1
                last = None
                for c in range(ord('E'), ord('G') + 1).__reversed__():
                    if isinstance(ws[chr(c) + str(index)].value, int):
                        last = c
                        break
                first = None
                if last:
                    for c in range(ord('E'), ord('G') + 1):
                        if isinstance(ws[chr(c) + str(index)].value, int):
                            first = c
                            break
                if first and last:
                    diff = ws[chr(last) + str(index)].value - ws[chr(first) + str(index)].value
                    ws['H' + str(index)] = diff
                else:
                    ws['H' + str(index)] = 'N/A'
                index += 1

        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
        for col, value in dims.items():
            ws.column_dimensions[col].width = int(value) + 2

    wb.save(filename='/tmp/lesfimi_hragogn.xlsx')
