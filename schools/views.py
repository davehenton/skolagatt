# -*- coding: utf-8 -*-
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic import ListView, CreateView, DetailView, UpdateView, DeleteView
from django.core.urlresolvers import reverse_lazy
from django.contrib.auth.mixins import UserPassesTestMixin
from django.contrib.auth.models import User
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

from itertools import chain
from uuid import uuid4
from datetime import datetime, date
from ast import literal_eval

import json
import xlrd
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.chart import AreaChart, BarChart, LineChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout

from kennitala import Kennitala

import common.mixins as common_mixins
from common.models import (
    Notification, School, Manager, Teacher, Student,
    StudentGroup, GroupSurvey, SurveyResult, SurveyLogin,
    ExampleSurveyQuestion, ExampleSurveyAnswer,
)
import common.util as common_util
import common.forms as cm_forms
import samraemd.models as s_models
from survey.models import (
    Survey, SurveyGradingTemplate, SurveyInputField,
    SurveyInputGroup, SurveyResource, SurveyTransformation,
    SurveyType
)

from supportandexception.models import (
    SupportResource)

import supportandexception.models as sae_models
from schools.tasks import save_example_survey_answers


def lesferill(request):
    return render(request, 'common/lesferill.html')


class NotificationCreate(UserPassesTestMixin, CreateView):
    model = Notification
    form_class = cm_forms.NotificationForm

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super(NotificationCreate, self).dispatch(request, *args, **kwargs)

    def test_func(self):
        return self.request.user.is_authenticated

    def post(self, *args, **kwargs):
        try:
            notification_type = self.request.POST.get('notification_type', None)
            notification_id = self.request.POST.get('notification_id', None)

            if notification_type and notification_id:
                Notification.objects.create(
                    user=self.request.user,
                    notification_type=notification_type,
                    notification_id=notification_id)

            return JsonResponse({'result': 'success'})
        except Exception as e:
            return JsonResponse({'result': 'error'})


class SchoolListing(ListView):
    model = School

    def get_context_data(self, **kwargs):
        context = super(SchoolListing, self).get_context_data(**kwargs)
        if self.request.user.is_authenticated:
            if self.request.user.is_superuser:
                context['school_list'] = common_util.slug_sort(
                    School.objects.all(), 'name')
            else:
                manager_schools = School.objects.filter(
                    managers=Manager.objects.filter(user=self.request.user)).all()
                teacher_schools = School.objects.filter(
                    teachers=Teacher.objects.filter(user=self.request.user)).all()
                context['school_list'] = list(
                    set(common_util.slug_sort(manager_schools | teacher_schools, 'name')))

        return context

    def get(self, *args, **kwargs):
        school_list = []
        if self.request.user.is_authenticated and not self.request.user.is_superuser:
            manager = Manager.objects.filter(user=self.request.user)
            teacher = Teacher.objects.filter(user=self.request.user)
            if manager.exists():
                manager = manager.first()
                school_list += manager.school_set.all()
            elif teacher.exists():
                teacher = teacher.first()
                school_list += teacher.school_set.all()
                if teacher.studentgroup_set.count() == 1:
                    studentgroup = teacher.studentgroup_set.first()
                    return redirect(reverse_lazy('schools:group_detail', kwargs={
                        'school_id': studentgroup.school.id,
                        'pk': studentgroup.id,
                    }))

            school_list = list(set(school_list))

        if len(school_list) == 1:
            return redirect(reverse_lazy('schools:school_detail', kwargs={'pk': school_list[0].id}))

        return super(SchoolListing, self).get(*args, **kwargs)


class SchoolDetail(common_mixins.SchoolEmployeeMixin, DetailView):
    model = School

    def get_context_data(self, **kwargs):
        context = super(SchoolDetail, self).get_context_data(**kwargs)
        context['studentgroup_list'] = common_util.slug_sort(
            self.object.studentgroup_set.all(), 'name')
        context['managers'] = common_util.slug_sort(
            self.object.managers.all(), 'name')
        context['teachers'] = common_util.slug_sort(
            self.object.teachers.all(), 'name')
        context['surveys'] = common_util.slug_sort(
            GroupSurvey.objects.filter(
                studentgroup__in=self.object.studentgroup_set.all()), 'survey')
        context['students'] = self.object.students.all()

        all_messages = common_util.get_messages()
        messages = []
        for message in all_messages:
            if message['file'] is not None:
                file_name = message['file'].split('/')[-1]
                file_name = file_name.encode('utf-8')
                message['file_name'] = file_name
            messages.append(message)

        if common_util.is_school_teacher(self.request, self.kwargs):
            context['messages'] = [
                message for message in messages if message['teachers_allow'] is True
            ]
        else:
            context['messages'] = messages

        return context


class SchoolCreate(common_mixins.SuperUserMixin, CreateView):
    model = School
    form_class = cm_forms.SchoolForm
    success_url = reverse_lazy('schools:school_listing')


class SchoolCreateImport(common_mixins.SchoolManagerMixin, CreateView):
    model = School
    form_class = cm_forms.SchoolForm
    template_name = "common/school_form_import.html"

    def get_context_data(self, **kwargs):
        # xxx will be available in the template as the related objects
        context = super(SchoolCreateImport, self).get_context_data(**kwargs)
        return context

    def post(self, *args, **kwargs):
        if(self.request.FILES):
            u_file = self.request.FILES['file'].name
            extension = u_file.split(".")[-1]
            ssn = self.request.POST.get('school_ssn')
            name = self.request.POST.get('school_name')
            title = self.request.POST.get('title')
            if title == 'yes':
                first = 1
            else:
                first = 0
            data = []
            errors = []
            if extension == 'xlsx':
                input_excel = self.request.FILES['file']
                book = xlrd.open_workbook(file_contents=input_excel.read())
                for sheetsnumber in range(book.nsheets):
                    sheet = book.sheet_by_index(sheetsnumber)
                    for row in range(first, sheet.nrows):
                        rowerrors = []
                        ssn_str = str(int(sheet.cell_value(row, int(ssn))))
                        kt = Kennitala(ssn_str)
                        if not kt.validate() or not kt.is_personal(ssn_str):
                            rowerrors.append({
                                'text': 'Kennitala ekki rétt slegin inn',
                                'row': row,
                            })
                        data.append({
                            'name': sheet.cell_value(row, int(name)),
                            'ssn': ssn_str,
                            'error': True if rowerrors else False,
                        })
                        if rowerrors:
                            errors += rowerrors
            common_util.store_import_data(self.request, 'school_create_import', data)

            return render(self.request, 'excel_verify_import.html', {
                'data': data,
                'errors': errors,
                'cancel_url': reverse_lazy('schools:school_listing'),
            })
        else:
            school_data = common_util.get_import_data(self.request, 'school_create_import')

            if school_data:
                # iterate through students, add them if they don't exist then add to school
                for school in school_data:
                    try:
                        School.objects.create(**school)
                    except:
                        pass  # student already exists

        return HttpResponseRedirect(self.get_success_url())

    def get_success_url(self):
        return reverse_lazy('schools:school_listing')


class SchoolUpdate(common_mixins.SuperUserMixin, UpdateView):
    model = School
    form_class = cm_forms.SchoolForm
    success_url = reverse_lazy('schools:school_listing')


class SchoolDelete(common_mixins.SuperUserMixin, DeleteView):
    model = School
    template_name = "schools/confirm_delete.html"
    success_url = reverse_lazy('schools:school_listing')


class ManagerListing(common_mixins.SchoolEmployeeMixin, ListView):
    model = Manager

    def get_context_data(self, **kwargs):
        context = super(ManagerListing, self).get_context_data(**kwargs)
        school = School.objects.get(pk=self.kwargs['school_id'])
        context['school'] = school
        context['managers'] = Manager.objects.filter(school=school)
        return context


class ManagerDetail(common_mixins.SchoolEmployeeMixin, DetailView):
    model = Manager

    def get_context_data(self, **kwargs):
        context = super(ManagerDetail, self).get_context_data(**kwargs)
        context['school'] = School.objects.get(pk=self.kwargs['school_id'])
        return context


class ManagerOverview(UserPassesTestMixin, DetailView):
    model = Manager

    def test_func(self):
        try:
            if self.request.user.id == Manager.objects.get(pk=self.kwargs['pk']).user.id:
                return True
        except:
            return False
        return False

    def get_context_data(self, **kwargs):
        context = super(ManagerOverview, self).get_context_data(**kwargs)
        context['school'] = School.objects.filter(
            managers=Manager.objects.filter(
                pk=self.kwargs['pk'])).first()
        return context


class ManagerCreate(common_mixins.SchoolManagerMixin, CreateView):
    model = Manager
    form_class = cm_forms.ManagerForm

    def get_context_data(self, **kwargs):
        context = super(ManagerCreate, self).get_context_data(**kwargs)
        context['school'] = School.objects.get(pk=self.kwargs['school_id'])
        return context

    def post(self, *args, **kwargs):
        self.object = Manager.objects.filter(
            user__username=self.request.POST.get('ssn')).first()
        if self.object:
            return HttpResponseRedirect(self.get_success_url())
        form = self.get_form()
        # make data mutable
        form.data = self.request.POST.copy()
        # set or create user
        user = User.objects.filter(username=self.request.POST.get('ssn'))
        if user:
            form.data['user'] = user.first().id
        else:
            new_user = User.objects.create(
                username=self.request.POST.get('ssn'),
                password=str(uuid4()))
            form.data['user'] = new_user.id
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    def get_success_url(self):
        try:
            school_id = self.kwargs['school_id']
            # Try getting the School object. If it doesn't exist it should return an error
            School.objects.get(pk=school_id).managers.add(self.object)
            return reverse_lazy(
                'schools:school_detail',
                kwargs={'pk': school_id})
        except:
            return reverse_lazy('schools:school_listing')

    def get_initial(self):
        school = get_object_or_404(School, pk=self.kwargs.get('school_id'))
        return {
            'school': school,
        }


class ManagerCreateImport(common_mixins.SchoolManagerMixin, CreateView):
    model = School
    form_class = cm_forms.SchoolForm
    template_name = "common/manager_form_import.html"

    def get_context_data(self, **kwargs):
        # xxx will be available in the template as the related objects
        context = super(ManagerCreateImport, self).get_context_data(**kwargs)
        return context

    def post(self, *args, **kwargs):
        if(self.request.FILES):
            u_file = self.request.FILES['file'].name
            extension = u_file.split(".")[-1]
            ssn = self.request.POST.get('ssn')
            name = self.request.POST.get('name')
            school_ssn = self.request.POST.get('school_ssn')
            title = self.request.POST.get('title')
            if title == 'yes':
                first = 1
            else:
                first = 0
            data = []
            errors = []
            if extension == 'xlsx':
                input_excel = self.request.FILES['file']
                book = xlrd.open_workbook(file_contents=input_excel.read())
                for sheetsnumber in range(book.nsheets):
                    sheet = book.sheet_by_index(sheetsnumber)
                    for row in range(first, sheet.nrows):
                        rowerrors = []
                        ssn_str = str(int(sheet.cell_value(row, int(ssn))))
                        school_ssn_str = str(int(sheet.cell_value(row, int(school_ssn))))

                        ssn_kt = Kennitala(ssn_str)
                        if not ssn_kt.validate() or not ssn_kt.is_personal(ssn_str):
                            rowerrors.append({
                                'text': 'Kennitala ekki rétt slegin inn',
                                'row': row,
                            })

                        school_ssn_kt = Kennitala(school_ssn_str)
                        if not school_ssn_kt.validate():
                            rowerrors.append({
                                'text': 'Kennitala skóla ekki rétt slegin inn',
                                'row': row
                            })

                        if not School.objects.filter(ssn=school_ssn_str).exists():
                            rowerrors.append({
                                'text': 'Skóli ekki til',
                                'row': row
                            })

                        data.append({
                            'name': sheet.cell_value(row, int(name)),
                            'ssn': ssn_str,
                            'school_ssn': school_ssn_str,
                            'error': True if rowerrors else False
                        })
                        if rowerrors:
                            errors += rowerrors
            common_util.store_import_data(self.request, 'manager_create_import', data)
            return render(self.request, 'excel_verify_import.html', {
                'data': data,
                'errors': errors,
                'cancel_url': reverse_lazy('schools:school_listing')
            })
        else:
            manager_data = common_util.get_import_data(self.request, 'manager_create_import')
            # iterate through managers, add them if they don't exist then add to school
            for manager in manager_data:
                try:
                    School.objects.get(
                        ssn=manager['school_ssn']
                    ).managers.add(**manager)
                except:
                    pass  # manager already exists

        return HttpResponseRedirect(self.get_success_url())

    def get_success_url(self):
        return reverse_lazy('schools:school_listing')


class ManagerUpdate(common_mixins.SchoolManagerMixin, UpdateView):
    model = Manager
    form_class = cm_forms.ManagerForm

    def get_context_data(self, **kwargs):
        context = super(ManagerUpdate, self).get_context_data(**kwargs)
        context['school'] = School.objects.get(pk=self.kwargs['school_id'])
        return context

    def post(self, request, **kwargs):
        # get Manager to be updated
        self.object = Manager.objects.get(pk=self.kwargs['pk'])
        # get user by ssn
        try:
            user = User.objects.get(username=self.request.POST.get('ssn'))
        except:
            user = User.objects.create(username=self.request.POST.get('ssn'), password=str(uuid4()))

        if self.object.user == user:
            # no changes
            pass
        else:
            request.POST = request.POST.copy()
            request.POST['user'] = user.id

        return super(ManagerUpdate, self).post(request, **kwargs)

    def get_success_url(self):
        try:
            school_id = self.kwargs['school_id']
            return reverse_lazy(
                'schools:school_detail',
                kwargs={'pk': school_id}
            )
        except:
            return reverse_lazy('schools:school_listing')


class ManagerDelete(common_mixins.SchoolManagerMixin, DeleteView):
    model = Manager
    template_name = "schools/confirm_delete.html"

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        # delete manager_school entry
        School.objects.get(
            pk=self.kwargs.get('school_id')
        ).managers.remove(self.object)
        return HttpResponseRedirect(self.get_success_url())

    def get_success_url(self):
        try:
            return reverse_lazy(
                'schools:school_detail',
                kwargs={'pk': self.kwargs['school_id']}
            )
        except:
            return reverse_lazy('schools:school_listing')


class TeacherListing(common_mixins.SchoolEmployeeMixin, ListView):
    model = Teacher

    def get_context_data(self, **kwargs):
        context = super(TeacherListing, self).get_context_data(**kwargs)
        school = School.objects.get(pk=self.kwargs['school_id'])
        context['school'] = school
        context['teachers'] = Teacher.objects.filter(school=school)
        context['groups'] = common_util.slug_sort(
            StudentGroup.objects.filter(school=school), 'name')
        return context


class TeacherDetail(common_mixins.SchoolEmployeeMixin, DetailView):
    model = Teacher

    def get_context_data(self, **kwargs):
        context = super(TeacherDetail, self).get_context_data(**kwargs)
        context['school'] = School.objects.get(pk=self.kwargs['school_id'])
        context['groups'] = common_util.slug_sort(
            StudentGroup.objects.filter(
                group_managers=Teacher.objects.filter(
                    user=self.request.user)
            ), 'name'
        )
        return context


class TeacherOverview(UserPassesTestMixin, DetailView):
    model = Teacher

    def test_func(self):
        try:
            if self.request.user.id == Teacher.objects.get(
                pk=self.kwargs['pk']
            ).user.id:
                return True
        except:
            return False
        return False

    def get_context_data(self, **kwargs):
        context = super(TeacherOverview, self).get_context_data(**kwargs)
        context['school'] = School.objects.filter(
            teachers=Teacher.objects.filter(
                pk=self.kwargs['pk']
            )
        ).first()
        return context


class TeacherCreate(common_mixins.SchoolManagerMixin, CreateView):
    model = Teacher
    form_class = cm_forms.TeacherForm

    def get_context_data(self, **kwargs):
        context = super(TeacherCreate, self).get_context_data(**kwargs)
        context['school'] = School.objects.get(pk=self.kwargs['school_id'])
        return context

    def post(self, *args, **kwargs):
        self.object = Teacher.objects.filter(
            ssn=self.request.POST.get('ssn')
        ).first()
        if self.object:
            return HttpResponseRedirect(self.get_success_url())
        form = self.get_form()
        # make data mutable
        form.data = self.request.POST.copy()
        # set or create user
        user = User.objects.filter(username=self.request.POST.get('ssn'))

        if user:
            form.data['user'] = user.first().id
        else:
            user = User.objects.create(
                username=self.request.POST.get('ssn'),
                password=str(uuid4()))
            form.data['user'] = user.id

        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    def get_success_url(self):
        try:
            school_id = self.kwargs['school_id']
            School.objects.get(pk=school_id).teachers.add(self.object)
            return reverse_lazy(
                'schools:school_detail',
                kwargs={'pk': school_id}
            )
        except:
            return reverse_lazy('schools:school_listing')


class TeacherCreateImport(common_mixins.SchoolManagerMixin, CreateView):
    model = Teacher
    form_class = cm_forms.TeacherForm
    template_name = "common/teacher_form_import.html"

    def get_context_data(self, **kwargs):
        # xxx will be available in the template as the related objects
        context = super(TeacherCreateImport, self).get_context_data(**kwargs)
        context['school'] = School.objects.get(pk=self.kwargs['school_id'])
        return context

    def post(self, *args, **kwargs):
        if(self.request.FILES):
            u_file = self.request.FILES['file'].name
            extension = u_file.split(".")[-1]
            ssn = self.request.POST.get('ssn')
            name = self.request.POST.get('name')
            title = self.request.POST.get('title')
            if title == 'yes':
                first = 1
            else:
                first = 0
            data = []
            errors = []

            if extension == 'xlsx':
                input_excel = self.request.FILES['file']
                book = xlrd.open_workbook(file_contents=input_excel.read())
                for sheetsnumber in range(book.nsheets):
                    sheet = book.sheet_by_index(sheetsnumber)
                    for row in range(first, sheet.nrows):
                        rowerrors = []
                        ssn_str = str(sheet.cell_value(row, int(ssn)))
                        ssn_kt = Kennitala(ssn_str)

                        if not ssn_kt.validate() or not ssn_kt.is_personal(ssn_kt):
                            rowerrors.append({
                                'text': 'Kennitala ekki rétt slegin inn',
                                'row': row,
                            })

                        data.append({
                            'name': str(sheet.cell_value(row, int(name))),
                            'ssn': ssn_str,
                            'error': True if rowerrors else False,
                        })

                        if rowerrors:
                            errors += rowerrors

            common_util.store_import_data(self.request, 'teacher_create_import', data)
            return render(
                self.request,
                'excel_verify_import.html',
                {
                    'data': data,
                    'errors': errors,
                    'cancel_url': reverse_lazy('schools:school_listing')
                }
            )
        else:
            teacher_data = common_util.get_import_data(self.request, 'teacher_create_import')
            # iterate through teachers, add them if they don't exist then add to school
            for teacher in teacher_data:
                try:
                    u, c = User.objects.get_or_create(username=teacher['ssn'])
                    t, c = Teacher.objects.get_or_create(
                        ssn=teacher['ssn'], name=teacher['name'], user=u)
                    School.objects.get(pk=self.kwargs['school_id']).teachers.add(t)
                except Exception as e:
                    pass  # teacher already exists
        return HttpResponseRedirect(self.get_success_url())

    def get_success_url(self):
        try:
            return reverse_lazy(
                'schools:school_detail',
                kwargs={'pk': self.kwargs['school_id']}
            )
        except:
            return reverse_lazy('schools:school_listing')


class TeacherUpdate(common_mixins.SchoolManagerMixin, UpdateView):
    model = Teacher
    form_class = cm_forms.TeacherForm

    def get_context_data(self, **kwargs):
        context = super(TeacherUpdate, self).get_context_data(**kwargs)
        context['school'] = School.objects.get(pk=self.kwargs['school_id'])
        return context

    def post(self, request, **kwargs):
        # get Teacher to be updated
        self.object = Teacher.objects.get(pk=self.kwargs['pk'])
        # get user by ssn
        try:
            user = User.objects.get(username=self.request.POST.get('ssn'))
        except:
            user = User.objects.create(username=self.request.POST.get('ssn'), password=str(uuid4()))

        if self.object.user == user:
            # no changes
            pass
        else:
            request.POST = request.POST.copy()
            request.POST['user'] = user.id

        return super(TeacherUpdate, self).post(request, **kwargs)

    def get_success_url(self):
        try:
            school_id = self.kwargs['school_id']
            return reverse_lazy(
                'schools:school_detail',
                kwargs={'pk': school_id}
            )
        except:
            return reverse_lazy('schools:school_listing')


class TeacherDelete(common_mixins.SchoolManagerMixin, DeleteView):
    model = Teacher
    template_name = "schools/confirm_delete.html"

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        # delete teacher_school entry
        School.objects.get(
            pk=self.kwargs.get('school_id')
        ).teachers.remove(self.object)
        return HttpResponseRedirect(self.get_success_url())

    def get_success_url(self):
        try:
            return reverse_lazy(
                'schools:school_detail',
                kwargs={'pk': self.kwargs['school_id']}
            )
        except:
            return reverse_lazy('schools:school_listing')


class StudentAdminSearch(common_mixins.SuperUserMixin, ListView):
    model = Student
    template_name = "common/student_search.html"


class StudentListing(common_mixins.SchoolEmployeeMixin, ListView):
    model = Student

    def get_context_data(self, **kwargs):
        context = super(StudentListing, self).get_context_data(**kwargs)
        school = School.objects.get(pk=self.kwargs['school_id'])

        context['school'] = school
        context['students'] = common_util.slug_sort(
            Student.objects.filter(school=school), 'name')
        context['students_outside_groups'] = common_util.slug_sort(
            Student.objects.filter(school=school).exclude(
                studentgroup__in=StudentGroup.objects.filter(school=school)
            ), 'name')
        return context


class StudentDetail(common_mixins.SchoolEmployeeMixin, DetailView):
    model = Student

    def get_context_data(self, **kwargs):
        # xxx will be available in the template as the related objects
        context = super(StudentDetail, self).get_context_data(**kwargs)

        if 'school_id' in self.kwargs:
            context['school'] = School.objects.get(pk=self.kwargs['school_id'])

        context['student_moreinfo'] = sae_models.StudentExceptionSupport.objects.filter(
            student=self.kwargs.get('pk')).get
        return context


def StudentNotes(request, school_id, pk):
    if request.method == 'POST':
        notes = request.POST.get('notes')
        if (sae_models.StudentExceptionSupport.objects.filter(
            student=Student.objects.get(pk=int(pk))
        ).exists()):
            sae_models.StudentExceptionSupport.objects.filter(
                student=Student.objects.get(pk=int(pk))).update(notes=notes)
        else:
            ses = sae_models.StudentExceptionSupport(notes=notes)
            ses.student = Student.objects.get(pk=int(pk))
            ses.save()
        return JsonResponse({"message": "success"})


class StudentCreateImport(common_mixins.SchoolManagerMixin, CreateView):
    model = Student
    form_class = cm_forms.StudentForm
    template_name = "common/student_form_import.html"

    def get_context_data(self, **kwargs):
        # xxx will be available in the template as the related objects
        context = super(StudentCreateImport, self).get_context_data(**kwargs)
        context['school'] = School.objects.get(pk=self.kwargs['school_id'])
        return context

    def post(self, *args, **kwargs):
        if(self.request.FILES):
            u_file = self.request.FILES['file'].name
            extension = u_file.split(".")[-1]
            ssn = self.request.POST.get('student_ssn')
            name = self.request.POST.get('student_name')
            title = self.request.POST.get('title')
            if title == 'yes':
                first = 1
            else:
                first = 0
            data = []
            errors = []
            if extension == 'xlsx':
                input_excel = self.request.FILES['file']
                book = xlrd.open_workbook(file_contents=input_excel.read())
                for sheetsnumber in range(book.nsheets):
                    rowerrors = []
                    sheet = book.sheet_by_index(sheetsnumber)
                    for row in range(first, sheet.nrows):
                        ssn_str = str(sheet.cell_value(row, int(ssn)))
                        ssn_kt = Kennitala(ssn_str)
                        if not ssn_kt.validate() or not ssn_kt.is_personal(ssn_str):
                            rowerrors.append({
                                'text': 'Kennitala rangt slegin inn',
                                row: row,
                            })
                        data.append({
                            'name': str(sheet.cell_value(row, int(name))).strip(),
                            'ssn': ssn_str,
                            'error': True if rowerrors else False,
                        })
                        if rowerrors:
                            errors += rowerrors
            common_util.store_import_data(self.request, 'student_create_import', data)
            return render(
                self.request,
                'excel_verify_import.html',
                {
                    'data': data,
                    'errors': errors,
                    'cancel_url': reverse_lazy('schools:school_detail', kwargs={'pk': self.kwargs['school_id']})
                }
            )
        else:
            student_data = common_util.get_import_data(self.request, 'student_create_import')
            school = School.objects.get(pk=self.kwargs['school_id'])
            # iterate through students, add them if they don't exist then add to school
            for data in student_data:
                student = None
                try:
                    student = Student.objects.create(**data)
                except:
                    student = Student.objects.get(ssn=data['ssn'])
                school.students.add(student)

        return HttpResponseRedirect(self.get_success_url())

    def get_success_url(self):
        try:
            school_id = self.kwargs['school_id']
            School.objects.get(pk=school_id).students.add(self.object)
            return reverse_lazy(
                'schools:school_detail',
                kwargs={'pk': school_id}
            )
        except:
            return reverse_lazy('schools:school_listing')


class StudentCreate(common_mixins.SchoolManagerMixin, CreateView):
    model = Student
    form_class = cm_forms.StudentForm

    def post(self, *args, **kwargs):
        self.object = Student.objects.filter(ssn=self.request.POST.get('ssn')).first()
        # if self.object:
        #     return HttpResponseRedirect(self.get_success_url())
        form = self.get_form()
        # make data mutable
        form.data = self.request.POST.copy()
        # set or create user
        user = User.objects.filter(username=self.request.POST.get('ssn'))
        if user:
            form.data['user'] = user
        else:
            form.data['user'] = User.objects.create(
                username=self.request.POST.get('ssn'),
                password=str(uuid4())
            )
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super(StudentCreate, self).get_context_data(**kwargs)
        context['school'] = School.objects.get(pk=self.kwargs['school_id'])
        return context

    def get_success_url(self):
        try:
            school_id = self.kwargs['school_id']
            School.objects.get(pk=school_id).students.add(self.object)
            return reverse_lazy(
                'schools:school_detail',
                kwargs={'pk': school_id}
            )
        except:
            return reverse_lazy('schools:school_listing')


class StudentUpdate(common_mixins.SchoolManagerMixin, UpdateView):
    model = Student
    form_class = cm_forms.StudentForm

    def get_context_data(self, **kwargs):
        context = super(StudentUpdate, self).get_context_data(**kwargs)

        if 'school_id' in self.kwargs:
            context['school'] = School.objects.get(pk=self.kwargs['school_id'])

        return context

    def get_success_url(self):
        try:
            school_id = self.kwargs['school_id']
            return reverse_lazy(
                'schools:school_detail',
                kwargs={'pk': school_id}
            )
        except:
            return reverse_lazy('schools:school_listing')


class StudentDelete(common_mixins.SchoolManagerMixin, DeleteView):
    model = Student
    template_name = "schools/confirm_delete.html"

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        # delete student_school entry
        if 'school_id' in self.kwargs:
            School.objects.get(
                pk=self.kwargs.get('school_id')
            ).students.remove(self.object)
            # remove student from all studentgroups in school
            for group in StudentGroup.objects.filter(school=self.kwargs.get('school_id')):
                group.students.remove(self.object)
        self.object.delete()
        return HttpResponseRedirect(self.get_success_url())

    def get_success_url(self):
        try:
            school_id = self.kwargs['school_id']
            return reverse_lazy(
                'schools:school_detail',
                kwargs={'pk': school_id}
            )
        except:
            return reverse_lazy('schools:school_listing')


class StudentGroupListing(common_mixins.SchoolEmployeeMixin, ListView):
    model = StudentGroup

    def get_context_data(self, **kwargs):
        context = super(StudentGroupListing, self).get_context_data(**kwargs)
        school = School.objects.get(pk=self.kwargs['school_id'])
        context['school'] = school
        context['studentgroups'] = common_util.slug_sort(
            StudentGroup.objects.filter(school=school), 'name')
        return context


class StudentGroupDetail(common_mixins.SchoolEmployeeMixin, DetailView):
    model = StudentGroup

    def get_context_data(self, **kwargs):
        # xxx will be available in the template as the related objects
        context = super(StudentGroupDetail, self).get_context_data(**kwargs)
        context['school'] = School.objects.get(pk=self.kwargs['school_id'])
        context['exceptions'] = sae_models.Exceptions.objects.all()
        context['supports'] = sae_models.SupportResource.objects.all()
        context['surveys'] = self.object.groupsurvey_set.filter(
            active_to__gte=datetime.now()
        )
        context['old_surveys'] = self.object.groupsurvey_set.filter(
            active_to__lt=datetime.now()
        )
        context['students'] = common_util.slug_sort(self.object.students.all(), 'name')
        context['teachers'] = common_util.slug_sort(self.object.group_managers.all(), 'name')
        context['samraemd'] = False
        samraemd_results = list(chain(
            s_models.SamraemdISLResult.objects.filter(
                student__in=self.object.students.all(),
                student_year=self.object.student_year,
            ),
            s_models.SamraemdMathResult.objects.filter(
                student__in=self.object.students.all(),
                student_year=self.object.student_year,
            ),
            s_models.SamraemdENSResult.objects.filter(
                student__in=self.object.students.all(),
                student_year=self.object.student_year,
            ),
        ))

        if samraemd_results:
            context['samraemd'] = True

        context['rawdata'] = ExampleSurveyAnswer.objects.filter(
            student__in=self.object.students.all(),
        ).exists()

        return context


class StudentGroupAdminListing(common_mixins.SuperUserMixin, ListView):
    model = StudentGroup
    template_name = "common/studentgroup_admin_list.html"

    def get_context_data(self, **kwargs):
        try:
            context = super(StudentGroupAdminListing, self).get_context_data(**kwargs)
            context['schools'] = common_util.slug_sort(School.objects.all(), 'name')
            context['surveys'] = GroupSurvey.objects.filter(
                survey__title=self.kwargs['survey_title']
            )
        except Exception as e:
            context['error'] = str(e)
        return context


class StudentGroupCreate(common_mixins.SchoolEmployeeMixin, CreateView):
    model = StudentGroup
    form_class = cm_forms.StudentGroupForm

    def get_context_data(self, **kwargs):
        # xxx will be available in the template as the related objects
        context = super(StudentGroupCreate, self).get_context_data(**kwargs)
        context['school'] = School.objects.get(pk=self.kwargs['school_id'])
        context['students'] = common_util.slug_sort(
            Student.objects.filter(school=self.kwargs['school_id']), 'name')
        context['teachers'] = common_util.slug_sort(
            Teacher.objects.filter(school=self.kwargs['school_id']), 'name')

        return context

    def post(self, request, *args, **kwargs):
        self.object = None
        form = self.get_form()
        # make data mutable
        form.data = self.request.POST.copy()
        form.data['school'] = School.objects.get(pk=self.kwargs['school_id']).pk
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    def get_success_url(self):
        try:
            school_id = self.kwargs['school_id']
            return reverse_lazy(
                'schools:school_detail',
                kwargs={'pk': school_id}
            )
        except:
            return reverse_lazy('schools:school_listing')


class StudentGroupUpdate(common_mixins.SchoolEmployeeMixin, UpdateView):
    model = StudentGroup
    form_class = cm_forms.StudentGroupForm

    def get_context_data(self, **kwargs):
        # xxx will be available in the template as the related objects
        context = super(StudentGroupUpdate, self).get_context_data(**kwargs)
        context['school'] = School.objects.get(pk=self.kwargs['school_id'])
        context['students'] = common_util.slug_sort(
            Student.objects.filter(school=self.kwargs['school_id']), 'name')
        context['teachers'] = common_util.slug_sort(
            Teacher.objects.filter(school=self.kwargs['school_id']), 'name')
        context['group_managers'] = Teacher.objects.filter(studentgroup=self.kwargs['pk'])
        return context

    def post(self, request, *args, **kwargs):
        self.object = None
        form = self.get_form()
        # make data mutable
        form.data = self.request.POST.copy()
        form.data['school'] = School.objects.get(pk=self.kwargs['school_id']).pk
        form = cm_forms.StudentGroupForm(
            form.data or None,
            instance=StudentGroup.objects.get(id=self.kwargs['pk'])
        )
        if form.is_valid():
            form.save()
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    def get_success_url(self):
        try:
            school_id = self.kwargs['school_id']
            return reverse_lazy(
                'schools:group_detail',
                kwargs={'school_id': school_id, 'pk': self.kwargs['pk']}
            )
        except:
            return reverse_lazy('schools:school_listing')


class StudentGroupDelete(common_mixins.SchoolEmployeeMixin, DeleteView):
    model = StudentGroup
    template_name = "schools/confirm_delete.html"

    def get_success_url(self):
        try:
            school_id = self.kwargs['school_id']
            return reverse_lazy(
                'schools:school_detail',
                kwargs={'pk': school_id}
            )
        except:
            return reverse_lazy('schools:school_listing')


class SurveyListing(common_mixins.SchoolEmployeeMixin, ListView):
    model = GroupSurvey

    def get_context_data(self, **kwargs):
        # xxx will be available in the template as the related objects
        context = super(SurveyListing, self).get_context_data(**kwargs)
        context['school'] = School.objects.get(pk=self.kwargs['school_id'])
        group = StudentGroup.objects.get(pk=self.kwargs['student_group'])
        context['studentgroup'] = group
        context['surveys'] = GroupSurvey.objects.filter(studentgroup=group)
        return context


class SurveyAdminListing(common_mixins.SuperUserMixin, ListView):
    model = GroupSurvey
    template_name = "common/survey_admin_list.html"

    def get_context_data(self, **kwargs):
        # xxx will be available in the template as the related objects
        try:
            context = super(SurveyAdminListing, self).get_context_data(**kwargs)
            context['surveys'] = Survey.objects.all()
        except Exception as e:
            context['error'] = str(e)
        return context


class SurveyDetail(common_mixins.SchoolEmployeeMixin, DetailView):
    model = GroupSurvey

    def get_context_data(self, **kwargs):
        # xxx will be available in the template as the related objects
        context = super(SurveyDetail, self).get_context_data(**kwargs)
        survey = Survey.objects.filter(pk=self.object.survey.id).first()
        survey_type = SurveyType.objects.filter(survey=self.object.survey.id).values('id')
        dic = survey_type[0]
        survey_type = dic['id']
        context['survey_details'] = survey
        context['survey_resources'] = SurveyResource.objects.filter(survey=survey)
        context['school'] = School.objects.get(pk=self.kwargs['school_id'])
        context['studentgroup'] = StudentGroup.objects.get(pk=self.kwargs['student_group'])
        transformation = SurveyTransformation.objects.filter(survey=survey)
        if transformation == []:
            transformation = -1

        try:
            context['students'] = self.object.studentgroup.students.all()
            context['expired'] = True if self.object.survey.active_to < date.today() else False
        except:
            context['students'] = []
            context['expired'] = False
        student_results = {}
        for student in context['students']:
            sr = SurveyResult.objects.filter(student=student, survey=self.object)
            if sr:
                r = literal_eval(sr.first().results)  # get student results
                try:
                    student_results[student] = common_util.calc_survey_results(
                        survey_identifier=self.object.survey.identifier,
                        click_values=literal_eval(r['click_values']),
                        input_values=r['input_values'],
                        student=student,
                        survey_type=survey_type,
                        transformation=transformation,
                    )
                except Exception as e:
                    student_results[student] = common_util.calc_survey_results(
                        survey_identifier=self.object.survey.identifier,
                        click_values=[],
                        input_values=r['input_values'],
                        student=student,
                        survey_type=survey_type,
                        transformation=transformation)

            else:
                student_results[student] = common_util.calc_survey_results(
                    survey_identifier=self.object.survey.identifier,
                    click_values=[],
                    input_values={},
                    student=student,
                    survey_type=survey_type)
        context['student_results'] = student_results
        context['field_types'] = ['text', 'number', 'text-list', 'number-list']
        return context


class SurveyResourcePrint(common_mixins.SchoolTeacherMixin, DetailView):
    model = SurveyResource
    template_name = "common/surveyresource_print.html"

    def get_context_data(self, **kwargs):
        # xxx will be available in the template as the related objects
        context = super(SurveyResourcePrint, self).get_context_data(**kwargs)
        context['school'] = School.objects.get(pk=self.kwargs['school_id'])
        context['studentgroup'] = StudentGroup.objects.get(pk=self.kwargs['student_group'])
        context['survey'] = GroupSurvey.objects.get(pk=self.kwargs['survey_id'])
        return context


class SurveyCreate(common_mixins.SchoolEmployeeMixin, CreateView):
    model = GroupSurvey
    form_class = cm_forms.SurveyForm

    def get_form(self):
        form = super(SurveyCreate, self).get_form(self.form_class)
        group_year = StudentGroup.objects.get(
            pk=self.kwargs['student_group']).student_year
        survey_set = Survey.objects.filter(student_year=group_year, active_to__gte=timezone.now().date())
        form.fields['survey'].queryset = survey_set
        return form

    def post(self, *args, **kwargs):
        self.object = None
        form = self.get_form()
        # make data mutable
        form.data = self.request.POST.copy()
        form.data['studentgroup'] = StudentGroup.objects.filter(
            pk=self.kwargs['student_group'])
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    def get_success_url(self):
        try:
            school_id = self.kwargs['school_id']
            student_group = self.kwargs['student_group']
            return reverse_lazy(
                'schools:group_detail',
                kwargs={'school_id': school_id, 'pk': student_group}
            )
        except:
            return reverse_lazy('schools:school_listing')


class SurveyUpdate(common_mixins.SchoolEmployeeMixin, UpdateView):
    model = GroupSurvey
    form_class = cm_forms.SurveyForm

    def get_form(self):
        form = super(SurveyUpdate, self).get_form(self.form_class)
        group_year = StudentGroup.objects.get(
            pk=self.kwargs['student_group']).student_year

        survey_set = Survey.objects.filter(student_year=group_year)
        form.fields['survey'].queryset = survey_set
        return form

    def get_context_data(self, **kwargs):
            # xxx will be available in the template as the related objects
        context = super(SurveyUpdate, self).get_context_data(**kwargs)
        try:
            context['survey_description'] = self.object.survey.description
        except:
            context['survey_description'] = "Veldu próf..."
        return context

    def get_success_url(self):
        try:
            school_id = self.kwargs['school_id']
            return reverse_lazy(
                'schools:school_detail',
                kwargs={'pk': school_id}
            )
        except:
            return reverse_lazy('schools:school_listing')


class SurveyDelete(common_mixins.SchoolEmployeeMixin, DeleteView):
    model = GroupSurvey
    template_name = "schools/confirm_delete.html"

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        success_url = self.get_success_url()
        # delete SurveyResults
        SurveyResult.objects.filter(survey=self.object).delete()
        self.object.delete()
        return HttpResponseRedirect(success_url)

    def get_success_url(self):
        try:
            school_id = self.kwargs['school_id']
            return reverse_lazy(
                'schools:school_detail',
                kwargs={'pk': school_id}
            )
        except:
            return reverse_lazy('schools:school_listing')


class SurveyResultCreate(common_mixins.SchoolEmployeeMixin, CreateView):
    model = SurveyResult
    form_class = cm_forms.SurveyResultForm

    def get_context_data(self, **kwargs):
        # xxx will be available in the template as the related objects
        context = super(SurveyResultCreate, self).get_context_data(**kwargs)
        context['data_result'] = "''"
        context['student'] = Student.objects.filter(pk=self.kwargs['student_id'])
        groupsurvey = GroupSurvey.objects.get(pk=self.kwargs['survey_id'])
        context['survey'] = groupsurvey
        try:
            survey = groupsurvey.survey
            if groupsurvey.is_expired():
                # survey is expired
                context['grading_template'] = ""
                context['info'] = ""
                context['input_fields'] = ""
            else:
                try:
                    grading_templates = SurveyGradingTemplate.objects.get(survey=survey)
                except:
                    grading_templates = []
                context['grading_template'] = grading_templates
                input_groups = SurveyInputGroup.objects.filter(survey=survey)
                i_gr = []
                for ig in input_groups:
                    i_gr.append({
                        'group': ig,
                        'inputs': SurveyInputField.objects.filter(input_group=ig)
                    })
                context['input_groups'] = i_gr
        except Exception as e:
            raise Exception("Ekki næst samband við prófagrunn")
        return context

    def post(self, *args, **kwargs):
        form = cm_forms.SurveyResultForm(self.request.POST)
        if form.is_valid():
            student = Student.objects.get(pk=self.kwargs['student_id'])
            survey = GroupSurvey.objects.get(pk=self.kwargs['survey_id'])
            survey_result = SurveyResult.objects.get_or_create(student=student, survey=survey)

            survey_result.reported_by = Teacher.objects.filter(user_id=self.request.user.pk).first()
            # extract data
            survey_result_data = {'click_values': [], 'input_values': {}}
            for k in self.request.POST:
                if k != 'csrfmiddlewaretoken':
                    if k == 'data_results[]':
                        survey_result_data['click_values'] = json.dumps(
                            self.request.POST.getlist('data_results[]')
                        )
                    else:
                        survey_result_data['input_values'][k] = self.request.POST[k]
            survey_result.results = json.dumps(survey_result_data)
            survey_result.save()
            return redirect(self.get_success_url())
        else:
            return self.form_invalid(form)

    def get_success_url(self):
        try:
            survey = GroupSurvey.objects.get(pk=self.kwargs['survey_id'])
            return reverse_lazy(
                'schools:survey_detail',
                kwargs={
                    'pk': self.kwargs['survey_id'],
                    'school_id': self.kwargs['school_id'],
                    'student_group': survey.studentgroup.id
                }
            )
        except:
            return reverse_lazy('schools:school_listing')


class SurveyResultUpdate(common_mixins.SchoolEmployeeMixin, UpdateView):
    model = SurveyResult
    form_class = cm_forms.SurveyResultForm

    def form_valid(self, form):
        survey_results = form.save(commit=False)
        survey_results.created_at = timezone.now()
        # extract data
        survey_results_data = {'click_values': [], 'input_values': {}}
        for k in self.request.POST:
            if k != 'csrfmiddlewaretoken':
                if k == 'data_results[]':
                    survey_results_data['click_values'] = json.dumps(
                        self.request.POST.getlist('data_results[]')
                    )
                else:
                    survey_results_data['input_values'][k] = self.request.POST[k]
        survey_results.results = json.dumps(survey_results_data)
        return super(SurveyResultUpdate, self).form_valid(form)

    def get_context_data(self, **kwargs):
        # xxx will be available in the template as the related objects
        context = super(SurveyResultUpdate, self).get_context_data(**kwargs)
        context['student'] = Student.objects.filter(pk=self.kwargs['student_id'])
        groupsurvey = GroupSurvey.objects.get(pk=self.kwargs['survey_id'])
        context['survey'] = GroupSurvey.objects.filter(pk=self.kwargs['survey_id'])
        context['data_result'] = json.loads(
            SurveyResult.objects.get(pk=self.kwargs['pk']).results
        ) or "''"
        survey = groupsurvey.survey
        try:
            grading_templates = SurveyGradingTemplate.objects.get(survey=survey)
        except:
            grading_templates = []
        context['grading_template'] = grading_templates
        input_groups = SurveyInputGroup.objects.filter(survey=survey)
        i_gr = []
        for ig in input_groups:
            i_gr.append({
                'group': ig,
                'inputs': SurveyInputField.objects.filter(input_group=ig)
            })
        context['input_groups'] = i_gr

        return context

    def get_success_url(self):
        try:
            survey = GroupSurvey.objects.get(pk=self.kwargs['survey_id'])
            return reverse_lazy(
                'schools:survey_detail',
                kwargs={
                    'pk': self.kwargs['survey_id'],
                    'school_id': self.kwargs['school_id'],
                    'student_group': survey.studentgroup.id
                }
            )
        except:
            return reverse_lazy('schools:school_listing')


class SurveyResultDelete(common_mixins.SchoolEmployeeMixin, DeleteView):
    model = SurveyResult
    template_name = "schools/confirm_delete.html"

    def get_success_url(self):
        try:
            school_id = self.kwargs['school_id']
            return reverse_lazy(
                'schools:school_detail',
                kwargs={'pk': school_id}
            )
        except:
            return reverse_lazy('schools:school_listing')


class SurveyLoginListing(common_mixins.SchoolEmployeeMixin, ListView):
    model = SurveyLogin

    def get_context_data(self, **kwargs):
        # xxx will be available in the template as the related objects
        context = super(SurveyLoginListing, self).get_context_data(**kwargs)
        school = School.objects.get(pk=self.kwargs['school_id'])
        context['school_id'] = school.id
        survey_list = SurveyLogin.objects.filter(
            student__in=Student.objects.filter(school=school)
        ).values('survey_id').distinct()
        survey_list = list(set([x['survey_id'].replace('_stuðningur', '') for x in survey_list]))
        context['survey_list'] = Survey.objects.filter(identifier__in=survey_list)
        print(context['survey_list'])
        return context


class SurveyLoginAdminListing(common_mixins.SuperUserMixin, ListView):
    model = SurveyLogin
    template_name = "common/surveylogin_admin_list.html"
    # Success url?

    def get_context_data(self, **kwargs):
        # xxx will be available in the template as the related objects
        context = super(SurveyLoginAdminListing, self).get_context_data(**kwargs)

        context['survey_login_list'] = SurveyLogin.objects.all(
        ).values('survey_id').distinct()
        return context


class SurveyLoginDetail(common_mixins.SchoolManagerMixin, DetailView):
    model = SurveyLogin

    def get_object(self, **kwargs):
        return SurveyLogin.objects.filter(survey_id=self.kwargs['survey_id']).first()

    def get_template_names(self, **kwargs):
        if 'print' in self.kwargs:
            if self.kwargs['print'] == 'listi':
                return ['surveylogin_detail_print_list.html']
            elif self.kwargs['print'] == 'einstaklingsblod':
                return ['surveylogin_detail_print_singles.html']
        return ['surveylogin_detail.html']

    def get_context_data(self, **kwargs):
        # xxx will be available in the template as the related objects
        context = super(SurveyLoginDetail, self).get_context_data(**kwargs)
        context['survey_id'] = self.kwargs['survey_id']
        identifier = self.kwargs['survey_id']
        if 'stuðningur' in identifier:
            survey_name = Survey.objects.filter(identifier=identifier[:len(identifier) - 11]).values('title')
        else:
            survey_name = Survey.objects.filter(identifier=identifier[:len(identifier)]).values('title')
        context['survey_name'] = survey_name[0]
        if 'school_id' in self.kwargs:
            school = School.objects.get(pk=self.kwargs['school_id'])

            context['survey_login_students'] = SurveyLogin.objects.filter(
                student__in=Student.objects.filter(school=school),
                survey_id=self.kwargs['survey_id']
            )
        else:
            context['survey_login_students'] = SurveyLogin.objects.filter(
                survey_id=self.kwargs['survey_id']
            )
        if 'íslenska' in self.kwargs['survey_id']:
            context['exam'] = '1'
        elif 'enska' in self.kwargs['survey_id']:
            context['exam'] = '2'
        elif 'stærðfræði' in self.kwargs['survey_id']:
            context['exam'] = '3'
        return context


class SurveyLoginCreate(common_mixins.SuperUserMixin, CreateView):
    model = SurveyLogin
    form_class = cm_forms.SurveyLoginForm
    template_name = "common/password_form_import.html"

    def post(self, *args, **kwargs):
        if(self.request.FILES):
            u_file = self.request.FILES['file'].name
            extension = u_file.split(".")[-1]
            ssn = self.request.POST.get('student_ssn')
            survey_id = self.request.POST.get('survey_id')
            password = self.request.POST.get('password')
            title = self.request.POST.get('title')

            if title == 'yes':
                first = 1
            else:
                first = 0

            data = []
            errors = []
            try:
                if extension == 'xlsx':
                    input_excel = self.request.FILES['file']
                    book = xlrd.open_workbook(file_contents=input_excel.read())
                    for sheetsnumber in range(book.nsheets):
                        sheet = book.sheet_by_index(sheetsnumber)
                        for row in range(first, sheet.nrows):
                            rowerrors = []
                            ssn_str = str(sheet.cell_value(row, int(ssn)))
                            ssn_kt = Kennitala(ssn_str)
                            if not ssn_kt.validate() or not ssn_kt.is_personal(ssn_str):
                                rowerrors.append({
                                    'text': 'Kennitala ekki rétt slegin inn',
                                    'row': row,
                                })

                            if not Student.objects.filter(ssn=ssn_str).exists():
                                rowerrors.append({
                                    'text': 'Nemandi ekki til',
                                    'row': row,
                                })
                            rowdata = {
                                'survey_id': str(sheet.cell_value(row, int(survey_id))),
                                'ssn': ssn_str,
                                'password': str(sheet.cell_value(row, int(password))),
                                'error': True if rowerrors else False
                            }
                            data.append(rowdata)
                            if rowerrors:
                                errors += rowerrors
                common_util.store_import_data(self.request, 'survey_login_create', data)
                return render(self.request, 'excel_verify_import.html', {
                    'data': data,
                    'errors': errors,
                    'cancel_url': reverse_lazy('schools:school_listing'),
                })
            except Exception as e:
                return render(
                    self.request,
                    'common/password_form_import.html',
                    {'error': 'Villa í skjali: "' + str(e) + ', lína: ' + str(row + 1)}
                )

        else:
            student_data = common_util.get_import_data(self.request, 'survey_login_create')
            # Iterate through the data
            # Add students if they don't exist then create a survey_login object
            for data in student_data:
                student = Student.objects.get(ssn=data['ssn'])  # student already exists

                # check if survey_login for student exists, create if not, otherwise update
                survey_login = SurveyLogin.objects.filter(
                    student=student,
                    survey_id=data['survey_id']
                )
                if survey_login:
                    survey_login.update(survey_code=data['password'])
                else:
                    survey_login = SurveyLogin.objects.create(
                        student=student,
                        survey_id=data['survey_id'],
                        survey_code=data['password']
                    )
        return redirect(self.get_success_url())

    def get_success_url(self):
        return reverse_lazy('schools:survey_login_admin_listing')


class SurveyLoginDelete(common_mixins.SuperUserMixin, DeleteView):
    model = SurveyLogin
    template_name = "schools/confirm_delete.html"

    def get_success_url(self):
        return reverse_lazy('schools:survey_login_admin_listing')

    def get_object(self):
        return SurveyLogin.objects.filter(survey_id=self.kwargs['survey_id'])


class AdminListing(common_mixins.SuperUserMixin, ListView):
    model = User
    template_name = "common/admin_list.html"

    def get_success_url(self):
        return reverse_lazy('schools:admin_listing')

    def get_context_data(self, **kwargs):
        # xxx will be available in the template as the related objects
        context = super(AdminListing, self).get_context_data(**kwargs)
        context['user_list'] = User.objects.filter(is_superuser=True)
        return context


class AdminCreate(common_mixins.SuperUserMixin, CreateView):
    model = User
    form_class = cm_forms.SuperUserForm
    template_name = "common/admin_form.html"

    def post(self, *args, **kwargs):
        username = self.request.POST.get('username')
        is_superuser = self.request.POST.get('is_superuser', False)
        if User.objects.filter(username=username).exists():
            if is_superuser:
                User.objects.filter(username=username).update(is_superuser=True)
        else:
            User.objects.create_user(
                username=username,
                is_superuser=is_superuser
            )
        return redirect(self.get_success_url())

    def get_success_url(self):
        return reverse_lazy('schools:admin_listing')


class AdminUpdate(common_mixins.SuperUserMixin, UpdateView):
    model = User
    form_class = cm_forms.SuperUserForm
    template_name = "common/admin_form.html"

    def get_success_url(self):
        return reverse_lazy('schools:admin_listing')


# Prófadæmi

class ExampleSurveyListing(common_mixins.SchoolTeacherMixin, ListView):
    model = ExampleSurveyAnswer
    template_name = "common/example_survey/listing.html"

    def get_context_data(self, **kwargs):
        # xxx will be available in the template as the related objects
        context = super(ExampleSurveyListing, self).get_context_data(**kwargs)
        school = School.objects.get(pk=self.kwargs['school_id'])
        context['school'] = school

        studentgroup = None
        if 'studentgroup_id' in self.kwargs:
            studentgroup = StudentGroup.objects.get(pk=self.kwargs['studentgroup_id'])

        samraemd_cats = ExampleSurveyQuestion.objects.all().values_list('quiz_type', flat=True).distinct()

        dates = []
        if studentgroup:
            dates = ExampleSurveyAnswer.objects.filter(
                student__in=studentgroup.students.all(),
                groupsurvey__isnull=True,
            ).values_list('date', flat=True).distinct()
        else:
            dates = ExampleSurveyAnswer.objects.filter(
                student__in=Student.objects.filter(school=school),
                groupsurvey__isnull=True,
            ).values_list('date', flat=True).distinct()

        samraemd = []
        for dte in dates:
            # Get all studentgroups for this school where a student in the
            # studentgroup has results in ExampleSurveyAnswer
            studentgroup_ids = []
            if studentgroup:
                studentgroup_ids = [studentgroup.id]
            else:
                studentgroup_ids = Student.objects.filter(
                    pk__in=ExampleSurveyAnswer.objects.filter(
                        student__in=Student.objects.filter(school=school),
                        date=dte,
                        groupsurvey__isnull=True,
                    ).values_list('student', flat=True).distinct()
                ).values_list('studentgroup', flat=True).distinct()
                studentgroup_ids = list(set(studentgroup_ids))
            studentgroups_list = []
            for studentgroup_id in studentgroup_ids:
                studentgroup = StudentGroup.objects.get(pk=studentgroup_id)
                students_list = []
                students = Student.objects.filter(studentgroup=studentgroup, school=school)

                for student in students:
                    quiz_type_list = ExampleSurveyQuestion.objects.filter(
                        pk__in=ExampleSurveyAnswer.objects.filter(
                            student=student,
                            date=dte,
                            groupsurvey__isnull=True,
                        ).values_list('question_id')
                    ).values_list('quiz_type', flat=True).distinct()
                    exceptions = []
                    if student.exceptions_set.exists():
                        excs = literal_eval(student.exceptions_set.first().exam)
                        if 1 in excs:
                            exceptions.append('ÍSL')
                        if 2 in excs:
                            exceptions.append('ENS')
                        if 3 in excs:
                            exceptions.append('STÆ')
                    students_list.append((student, quiz_type_list, exceptions))
                studentgroups_list.append((studentgroup, students_list))
            samraemd.append((dte, studentgroups_list))

        print(samraemd)
        surveys = []
        groupsurveys = GroupSurvey.objects.filter(
            pk__in=ExampleSurveyAnswer.objects.filter(
                student__in=Student.objects.filter(school=school)
            ).values_list('groupsurvey', flat=True).distinct().order_by('-id')
        ).order_by('-active_to').all()

        for groupsurvey in groupsurveys:
            students = groupsurvey.studentgroup.students.all()

            students_list = []
            for student in students:
                quiz_type_list = ExampleSurveyQuestion.objects.filter(
                    pk__in=ExampleSurveyAnswer.objects.filter(
                        student=student,
                        groupsurvey=groupsurvey,
                    ).values_list('question_id')
                ).values_list('quiz_type', flat=True).distinct()
                students_list.append((student, quiz_type_list))
            surveys.append((groupsurvey, students_list))

        context['surveys'] = surveys
        context['samraemd'] = samraemd
        context['samraemd_cats'] = samraemd_cats
        return context


class ExampleSurveyGSDetail(common_mixins.SchoolManagerMixin, ListView):
    model = ExampleSurveyAnswer
    template_name = "common/example_survey/survey_detail.html"

    def get_context_data(self, **kwargs):
        # xxx will be available in the template as the related objects
        context = super(ExampleSurveyGSDetail, self).get_context_data(**kwargs)

        school = School.objects.get(pk=self.kwargs['school_id'])
        student = Student.objects.get(pk=self.kwargs['student_id'])
        groupsurvey = GroupSurvey.objects.get(pk=self.kwargs['groupsurvey_id'])

        answers = ExampleSurveyAnswer.objects.filter(
            student=student,
            groupsurvey=groupsurvey,
        ).order_by('question__category', '?').all()

        context['answers'] = answers
        context['school'] = school
        context['student'] = student
        context['groupsurvey'] = groupsurvey
        context['studentgroup'] = groupsurvey.studentgroup

        return context


class ExampleSurveySamraemdDetail(common_mixins.SchoolTeacherMixin, ListView):
    model = ExampleSurveyAnswer

    def get_template_names(self, **kwargs):
        if 'print' in self.request.path:
            return ['common/example_survey/samraemd_detail_print.html']
        return ['common/example_survey/samraemd_detail.html']

    def _get_student_answers_list(self, student, quiz_type, year):
        answers_list = student.examplesurveyanswer_set.filter(
            groupsurvey__isnull=True,
            date__year=year,
            question__quiz_type=quiz_type,
        ).order_by('question__category').all()

        return answers_list

    def get_context_data(self, **kwargs):
        # xxx will be available in the template as the related objects
        context = super(ExampleSurveySamraemdDetail, self).get_context_data(**kwargs)

        school = School.objects.get(pk=self.kwargs['school_id'])
        year = self.kwargs['year']
        quiz_type = self.kwargs['quiz_type']
        student_answers = []
        if 'student_id' in self.kwargs:
            student = Student.objects.get(pk=self.kwargs['student_id'])
            answers = self._get_student_answers_list(student, quiz_type, year)
            student_answers.append((student, answers))
        else:
            studentgroup = StudentGroup.objects.get(pk=self.kwargs['studentgroup_id'])
            students = Student.objects.filter(studentgroup=studentgroup, school=school).all()

            for student in students:
                answers = self._get_student_answers_list(student, quiz_type, year)
                if not answers:
                    continue
                student_answers.append((student, answers))

        context['student_answers'] = student_answers
        context['school'] = school
        context['year'] = year
        context['quiz_type'] = quiz_type
        context['groupsurvey'] = None
        context['studentgroup'] = None

        return context


class ExampleSurveyQuestionAdminListing(common_mixins.SuperUserMixin, ListView):
    model = ExampleSurveyQuestion
    template_name = "common/example_survey/question_admin_list.html"
    # Success url?

    def get_context_data(self, **kwargs):
        context = super(ExampleSurveyQuestionAdminListing, self).get_context_data(**kwargs)

        context['questions'] = ExampleSurveyQuestion.objects.all()
        return context


class ExampleSurveyQuestionAdminDetail(common_mixins.SuperUserMixin, DetailView):
    model = ExampleSurveyQuestion
    template_name = "common/example_survey/question_admin_detail.html"

    def get_context_data(self, **kwargs):
        context = super(ExampleSurveyQuestionAdminDetail, self).get_context_data(**kwargs)

        question = ExampleSurveyQuestion.objects.get(pk=self.kwargs['pk'])
        context['question'] = question

        return context


class ExampleSurveyQuestionAdminCreate(common_mixins.SuperUserMixin, CreateView):
    model = ExampleSurveyQuestion
    form_class = cm_forms.ExampleSurveyQuestionForm
    template_name = "common/example_survey/question_admin_create.html"

    def form_valid(self, form):
        survey = form.save(commit=False)
        survey.created_by = self.request.user
        return super(ExampleSurveyQuestionAdminCreate, self).form_valid(form)

    def get_success_url(self):
        return reverse_lazy('schools:example_survey_question_admin_listing')


class ExampleSurveyQuestionAdminUpdate(common_mixins.SuperUserMixin, UpdateView):
    model = ExampleSurveyQuestion
    form_class = cm_forms.ExampleSurveyQuestionForm
    template_name = "common/example_survey/question_admin_create.html"

    def get_success_url(self):
        return reverse_lazy('schools:example_survey_question_admin_listing')


class ExampleSurveyQuestionAdminDelete(common_mixins.SuperUserMixin, DeleteView):
    model = ExampleSurveyQuestion
    template_name = "schools/confirm_delete.html"

    def get_success_url(self):
        return reverse_lazy('schools:example_survey_question_admin_listing')

    def get_object(self):
        return ExampleSurveyQuestion.objects.get(pk=self.kwargs['pk'])


class ExampleSurveyAnswerAdminListing(common_mixins.SuperUserMixin, ListView):
    model = ExampleSurveyAnswer
    template_name = "common/example_survey/answer_admin_list.html"
    # Success url?

    def get_context_data(self, **kwargs):
        context = super(ExampleSurveyAnswerAdminListing, self).get_context_data(**kwargs)

        if 'cancel_import' in self.request.GET:
            print("Cancel import, removing session data")
            common_util.cancel_import_data(self.request, 'example_survey_answer_admin_import')

        if 'exam_code' in self.kwargs:
            answers = ExampleSurveyAnswer.objects.filter(exam_code=self.kwargs['exam_code']).all()
            paginator = Paginator(answers, 25)
            page = self.request.GET.get('page')
            try:
                answers = paginator.page(page)
            except PageNotAnInteger:
                answers = paginator.page(1)
            except EmptyPage:
                answers = paginator.page(paginator.num_pages)
            context['answers'] = answers
        else:
            exam_codes = ExampleSurveyAnswer.objects.values_list('exam_code', flat=True).distinct()
            context['exam_codes'] = exam_codes

        context['jobs'] = common_util.get_celery_jobs('save_example_survey_answers')

        return context


class ExampleSurveyAnswerAdminDetail(common_mixins.SuperUserMixin, DetailView):
    model = ExampleSurveyQuestion
    template_name = "common/example_survey/answer_admin_detail.html"

    def get_context_data(self, **kwargs):
        context = super(ExampleSurveyAnswerAdminDetail, self).get_context_data(**kwargs)

        context['answer'] = ExampleSurveyAnswer.objects.get(pk=self.kwargs['pk'])
        return context


class ExampleSurveyAnswerAdminImport(common_mixins.SuperUserMixin, CreateView):
    model = ExampleSurveyQuestion
    form_class = cm_forms.ExampleSurveyAnswerForm
    template_name = "common/example_survey/answer_form_import.html"

    def post(self, *args, **kwargs):
        if(self.request.FILES):
            u_file = self.request.FILES['file'].name
            extension = u_file.split(".")[-1]

            ssn = int(self.request.POST.get('ssn'))
            quickcode = int(self.request.POST.get('quickcode'))
            answer = int(self.request.POST.get('answer'))
            title = self.request.POST.get('title')

            exam_date = self.request.POST.get('exam_date').strip()
            survey_identifier = self.request.POST.get('survey_identifier')

            if title == 'yes':
                first = 1
            else:
                first = 0

            data = []
            errors = []

            try:
                if extension == 'xlsx':
                    input_excel = self.request.FILES['file']
                    book = xlrd.open_workbook(file_contents=input_excel.read())
                    ssn_cache = {}
                    quickcode_cache = {}
                    for sheetsnumber in range(book.nsheets):
                        sheet = book.sheet_by_index(sheetsnumber)

                        if not survey_identifier:
                            errors.append({'text': 'Enginn prófkóði', 'row': 0})
                        else:
                            survey = Survey.objects.filter(identifier=survey_identifier).first()
                            if survey:
                                ssn_str = str(sheet.cell_value(1, ssn))
                                student = Student.objects.filter(ssn=ssn_str).first()
                                if student:
                                    groupsurvey = GroupSurvey.objects.filter(
                                        survey=survey,
                                        studentgroup=student.studentgroup,
                                    ).first()
                                    if groupsurvey and ExampleSurveyAnswer.objects.filter(survey=survey).exists():
                                        errors.append({
                                            'text': 'Svör þegar til fyrir þennan kóða, öll ný svör bætast við',
                                            'row': 0,
                                        })

                            else:
                                if ExampleSurveyAnswer.objects.filter(exam_code=survey_identifier).exists():
                                    errors.append({
                                        'text': 'Svör þegar til fyrir þennan kóða, öll ný svör bætast við',
                                        'row': 0,
                                    })

                        for row in range(first, sheet.nrows):
                            ssn_str = str(sheet.cell_value(row, ssn))
                            quickcode_str = str(sheet.cell_value(row, quickcode))
                            answer_str = str(sheet.cell_value(row, answer))

                            rowerrors = []

                            if ssn_str not in ssn_cache.keys():
                                ssn_cache[ssn_str] = Student.objects.filter(ssn=ssn_str).exists()
                                if not ssn_cache[ssn_str]:
                                    rowerrors.append({
                                        'text': 'Nemandi {} er ekki til'.format(ssn_str),
                                        'row': row,
                                    })

                            if quickcode_str not in quickcode_cache.keys():
                                quickcode_cache[quickcode_str] = ExampleSurveyQuestion.objects.filter(
                                    quickcode=quickcode_str).exists()
                                if not quickcode_cache[quickcode_str]:
                                    rowerrors.append({
                                        'text': 'Prófdæmi {} er ekki til'.format(quickcode_str),
                                        'row': row,
                                    })

                            if answer_str not in ['1', '0']:
                                rowerrors.append({
                                    'text': 'Get ekki notað svar {}'.format(answer_str),
                                    'row': row,
                                })

                            data.append({
                                'ssn': ssn_str,
                                'quickcode': quickcode_str,
                                'survey_identifier': survey_identifier,
                                'exam_date': exam_date,
                                'answer': answer_str,
                                'error': True if rowerrors else False,
                            })
                            if rowerrors:
                                errors += rowerrors
                    common_util.store_import_data(self.request, 'example_survey_answer_admin_import', data)
                return render(self.request, 'excel_verify_import.html', {
                    'data': [],
                    'errors': errors,
                    'cancel_url': reverse_lazy('schools:example_survey_answer_admin_listing') + "?cancel_import"
                })
            except Exception as e:
                return render(
                    self.request,
                    'common/example_survey/answer_form_import.html',
                    {'error': 'Villa í skjali: "' + str(e) + ', lína: ' + str(row + 1)}
                )
        else:
            newdata = common_util.get_import_data(self.request, 'example_survey_answer_admin_import')
            if newdata:
                print("Calling save_example_survey_answers for import")
                save_example_survey_answers.delay(newdata)
            else:
                print("Import likely timed out")
        return redirect(self.get_success_url())

    def get_success_url(self):
        return reverse_lazy('schools:example_survey_answer_admin_listing')


class ExampleSurveyAnswerAdminDelete(common_mixins.SuperUserMixin, DeleteView):
    model = ExampleSurveyAnswer
    template_name = "schools/confirm_delete.html"

    def get_success_url(self):
        return reverse_lazy('schools:example_survey_answer_admin_listing')

    def get_object(self):
        if 'pk' in self.kwargs:
            return ExampleSurveyAnswer.objects.get(pk=self.kwargs['pk'])
        return ExampleSurveyAnswer.objects.filter(
            exam_code=self.kwargs['exam_code'],
        )


def group_admin_listing_excel(request, survey_title):
    surveys = GroupSurvey.objects.filter(survey__title=survey_title)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=samræmdupróf.xlsx'
    wb = openpyxl.Workbook()

    ws = wb.get_active_sheet()
    ws.title = 'samræmdupróf'
    ws['A1'] = 'FIRST'
    ws['B1'] = 'LAST'
    ws['C1'] = 'RESULTS EMAIL'
    ws['D1'] = 'EXTERNAL ID'
    ws['E1'] = 'CODE EMAIL'
    ws['F1'] = 'TEST NAME'
    ws['G1'] = 'GROUP PATH'
    ws['H1'] = 'EXTRA TIME'
    ws['J1'] = 'BEKKUR (TAKA ÚT)'

    index = 2
    for survey in surveys:
        if survey.studentgroup:
            for student in survey.studentgroup.students.all():
                names_list = student.name.strip().split()
                pos = len(names_list) - 1
                ws.cell('A' + str(index)).value = ' '.join(names_list[:pos])
                ws.cell('B' + str(index)).value = ' '.join(names_list[pos:])
                ws.cell('D' + str(index)).value = student.ssn
                ws.cell('F' + str(index)).value = survey.survey.identifier
                ws.cell('H' + str(index)).value = 'N'
                supports = SupportResource.objects.filter(student=student.id)
                for support in supports:
                    if 'fyrri' in survey_title:
                        if '1' in support.longer_time or '2' in support.longer_time or \
                                '1' in support.reading_assistance or '2' in support.reading_assistance:
                            ws.cell('H' + str(index)).value = 'Y'
                    elif 'seinni' in survey_title:
                        if '2' in support.longer_time or '3' in support.longer_time or \
                                '2' in support.reading_assistance or '3' in support.reading_assistance:
                            ws.cell('H' + str(index)).value = 'Y'

                ws.cell('J' + str(index)).value = str(StudentGroup.objects.filter(students=student.id)[:1].get())
                index += 1

    wb.save(response)

    return response


def group_admin_attendance_excel(request, survey_title):

    surveys = GroupSurvey.objects.filter(survey__title=survey_title)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=samræmdupróf.xlsx'

    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = 'samræmdupróf'

    ws['A1'] = 'Skóli'
    ws['B1'] = 'Kennitala sḱóla'
    ws['C1'] = 'Nemandi'
    ws['D1'] = 'Kennitala'
    ws['E1'] = 'Bekkur'
    ws['F1'] = 'Mættur'
    ws['G1'] = 'Undanþága'
    ws['H1'] = 'Veikur'
    ws['I1'] = 'Fjarverandi'
    ws['J1'] = 'Stuðningur í íslensku'
    ws['K1'] = 'Stuðningur í Stærðfræði'

    index = 2
    for survey in surveys:
        if survey.studentgroup:
            for student in survey.studentgroup.students.all():
                ws.cell('A' + str(index)).value = survey.studentgroup.school.name
                ws.cell('B' + str(index)).value = survey.studentgroup.school.ssn
                ws.cell('C' + str(index)).value = student.name
                ws.cell('D' + str(index)).value = student.ssn
                ws.cell('E' + str(index)).value = survey.studentgroup.name

                sr = SurveyResult.objects.filter(student=student, survey=survey)
                student_results = ""
                if sr:
                    r = literal_eval(sr.first().results)  # get student results
                    try:
                        student_results = int(r['click_values'][2])
                    except Exception as e:
                        print(e)
                else:
                    student_results = ''

                if student_results != '':
                    ws.cell(row=index, column=student_results + 6).value = "1"

                support = sae_models.SupportResource.objects.filter(student=student)

                if support:
                    support_available = []
                    support_available += literal_eval(support.first().reading_assistance)
                    support_available += literal_eval(support.first().interpretation)
                    support_available += literal_eval(support.first().longer_time)
                    if 1 in support_available:
                        ws.cell(row=index, column=10).value = "1"
                    if 3 in support_available:
                        ws.cell(row=index, column=11).value = "1"

                index += 1

    wb.save(response)

    return response


def survey_detail_excel(request, school_id, student_group, pk):
    studentgroup = StudentGroup.objects.get(pk=student_group)
    survey = GroupSurvey.objects.filter(pk=pk).first()

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Niðurstöður - {}.xlsx'.format(survey.survey.title)
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()

    identifier = survey.survey.identifier

    if identifier.startswith('01b_LTL_'):
        ws.title = 'Lesskimun fyrir fyrsta bekk'

        ws['A1'] = 'Kennitala'
        ws['B1'] = 'Nafn'
        ws['C1'] = 'Hljóð'
        ws['D1'] = 'Mál'
        ws['E1'] = 'Stafir'

        index = 2
        if studentgroup:
            for student in studentgroup.students.all():
                ws.cell('A' + str(index)).value = student.ssn
                ws.cell('B' + str(index)).value = student.name
                sr = SurveyResult.objects.filter(student=student, survey=survey)
                if sr:
                    r = literal_eval(sr.first().results)  # get student results
                    survey_student_result = common_util.calc_survey_results(
                        identifier,
                        [],
                        r['input_values'],
                        student
                    )
                    ws.cell('C' + str(index)).value = survey_student_result[0]
                    ws.cell('D' + str(index)).value = survey_student_result[1]
                    ws.cell('E' + str(index)).value = survey_student_result[2]
                else:
                    ws.cell('C' + str(index)).value = 'Vantar gögn'
                    ws.cell('D' + str(index)).value = 'Vantar gögn'
                    ws.cell('E' + str(index)).value = 'Vantar gögn'
                index += 1
    elif identifier.endswith('_LF_jan17'):
        ws.title = 'Niðurstöður'
        ws['A1'] = 'Kennitala'
        ws['B1'] = 'Nafn'
        ws['C1'] = 'September'
        ws['D1'] = 'Janúar'
        ws['E1'] = 'Mismunur'

        ref_values = {
            '1': (20, 55, 75),
            '2': (40, 85, 100),
            '3': (55, 100, 120),
            '4': (80, 120, 145),
            '5': (90, 140, 160),
            '6': (105, 155, 175),
            '7': (120, 165, 190),
            '8': (130, 180, 210),
            '9': (140, 180, 210),
            '10': (145, 180, 210),
        }

        index = 2
        survey_type = SurveyType.objects.filter(survey=survey.survey.id).values('id')
        dic = survey_type[0]
        survey_type = dic['id']
        jan_transformation = SurveyTransformation.objects.filter(survey=survey.survey)

        datapoints = []
        datapoints.append(('Nafn', 'September', 'Janúar', '90% viðmið', '50% viðmið', '25% viðmið'))

        if studentgroup:
            sept_identifier = "{}b_LF_sept".format(studentgroup.student_year)
            sept_survey = Survey.objects.filter(identifier=sept_identifier).first()
            sept_transformation = SurveyTransformation.objects.filter(survey=sept_survey)
            sept_gs = GroupSurvey.objects.filter(survey=sept_survey, studentgroup=survey.studentgroup).first()

            for student in studentgroup.students.all():
                value_jan = 'Vantar gögn'
                value_sept = 'Vantar gögn'
                ws.cell('A' + str(index)).value = student.ssn
                ws.cell('B' + str(index)).value = student.name
                sy = studentgroup.student_year
                sr = SurveyResult.objects.filter(student=student, survey=survey)
                if sr:
                    r = literal_eval(sr.first().results)  # get student results
                    try:
                        click_values = literal_eval(r['click_values'])
                    except:
                        click_values = []
                    survey_student_result = common_util.calc_survey_results(
                        survey_identifier=identifier,
                        click_values=click_values,
                        input_values=r['input_values'],
                        student=student,
                        survey_type=survey_type,
                        transformation=jan_transformation,
                    )
                    value_jan = survey_student_result[0]

                sr = SurveyResult.objects.filter(student=student, survey=sept_gs)
                if sr:
                    r = literal_eval(sr.first().results)  # get student results
                    try:
                        click_values = literal_eval(r['click_values'])
                    except:
                        click_values = []
                    survey_student_result = common_util.calc_survey_results(
                        survey_identifier=sept_identifier,
                        click_values=click_values,
                        input_values=r['input_values'],
                        student=student,
                        survey_type=survey_type,
                        transformation=sept_transformation,
                    )
                    value_sept = survey_student_result[0]

                ws.cell('C' + str(index)).value = value_sept
                ws.cell('D' + str(index)).value = value_jan
                if isinstance(value_sept, int) and isinstance(value_jan, int):
                    cur_cell = ws.cell('E' + str(index))
                    diff = value_jan - value_sept

                    # Student year
                    if int(sy) <= 3:
                        if diff <= 7:
                            cur_cell.fill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type="solid")

                        elif diff <= 20:
                            cur_cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type="solid")
                        else:
                            cur_cell.fill = PatternFill(start_color='00ff00', end_color='00ff00', fill_type="solid")
                    else:
                        if diff < 0:
                            cur_cell.fill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type="solid")

                        elif diff <= 10:
                            cur_cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type="solid")
                        else:
                            cur_cell.fill = PatternFill(start_color='00ff00', end_color='00ff00', fill_type="solid")

                    cur_cell.value = diff

                if not isinstance(value_sept, int):
                    value_sept = 0
                if not isinstance(value_jan, int):
                    value_jan = 0

                datapoints.append((student.name.split()[0], value_sept, value_jan, ref_values[
                                  sy][0], ref_values[sy][1], ref_values[sy][2]))

                index += 1

        # Fix column widths
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
        for col, value in dims.items():
            ws.column_dimensions[col].width = int(value) + 2

        # Add legend
        index += 2
        ws.cell('A' + str(index)).fill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')
        ws.cell('B' + str(index)).value = 'Lítil framvinda miðað við aldur'
        ws.merge_cells('B' + str(index) + ':E' + str(index))
        index += 1
        ws.cell('A' + str(index)).fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type='solid')
        ws.cell('B' + str(index)).value = 'Fremur lítil framvinda'
        ws.merge_cells('B' + str(index) + ':E' + str(index))
        index += 1
        ws.cell('A' + str(index)).fill = PatternFill(start_color='00ff00', end_color='00ff00', fill_type='solid')
        ws.cell('B' + str(index)).value = 'Góð framvinda'
        ws.merge_cells('B' + str(index) + ':E' + str(index))

        # Add bar chart
        ws_bar = wb.create_sheet(title="Súlurit")
        for datapoint in datapoints:
            ws_bar.append(datapoint)
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Niðurstöður úr Lesfimi, sept 2016 - jan 2017"
        chart.y_axis.title = "Orð á mínútu"
        chart.x_axis.title = "Nemandi"
        data = Reference(ws_bar, min_col=2, min_row=1, max_row=len(datapoints), max_col=3)
        cats = Reference(ws_bar, min_col=1, min_row=2, max_row=len(datapoints))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        chart.width = 40
        chart.height = 20

        # Add reference lines
        lchart = LineChart()
        ldata = Reference(ws_bar, min_col=4, max_col=6, min_row=1, max_row=len(datapoints) + 1)
        lchart.width = 40
        lchart.height = 20

        lchart.layout = Layout(
            ManualLayout(
                xMode="edge",
                yMode="edge",
            )
        )
        lchart.add_data(ldata, titles_from_data=True)
        chart += lchart

        for idx in range(1, len(datapoints) + 1):
            ws_bar.row_dimensions[idx].hidden = True

        ws_bar.add_chart(chart, "A" + str(len(datapoints) + 1))

    wb.save(response)

    return response


def lesfimi_excel_for_principals(request, pk):
    ref_values = {
        1: (20, 55, 75),
        2: (40, 85, 100),
        3: (55, 100, 120),
        4: (80, 120, 145),
        5: (90, 140, 160),
        6: (105, 155, 175),
        7: (120, 165, 190),
        8: (130, 180, 210),
        9: (140, 180, 210),
        10: (145, 180, 210),
    }

    # Get the school object
    school = School.objects.get(pk=pk)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Lesfimi {}.xlsx'.format(school.name)

    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    wb.remove_sheet(ws)

    tests = (
        ('b{}_LF_jan17', 'Janúar 2017'),
        ('{}b_LF_sept', 'September 2016'),
    )
    for test in tests:
        identifier = test[0]
        title = test[1]
        ws = wb.create_sheet(title=title)
        ws['A1'] = 'Árgangur'
        ws['B1'] = 'Fjöldi nemenda'
        ws['C1'] = 'Fjöldi sem þreytti próf'
        ws['D1'] = 'Hlutfall sem þreytti próf'
        ws['E1'] = 'Hlutfall sem nær 90% viðmiðum'
        ws['F1'] = 'Hlutfall sem nær 50% viðmiðum'
        ws['G1'] = 'Hlutfall sem nær 25% viðmiðum'
        index = 2
        errors = []
        for year in range(1, 11):
            ws['A' + str(index)] = year
            survey = Survey.objects.filter(identifier=identifier.format(year)).first()
            survey_type = SurveyType.objects.filter(survey=survey.id).values('id')
            dic = survey_type[0]
            survey_type = dic['id']
            transformation = SurveyTransformation.objects.filter(survey=survey)
            studentgroups = StudentGroup.objects.filter(school=school, student_year=year).all()
            this_year_result = {
                'students': 0,
                'students_who_took_test': 0,
                'students_over_25pct': 0,
                'students_over_50pct': 0,
                'students_over_90pct': 0,
            }
            for studentgroup in studentgroups:
                this_year_result['students'] += studentgroup.students.all().count()
                groupsurveys = GroupSurvey.objects.filter(studentgroup=studentgroup, survey=survey)
                if groupsurveys.all().count() > 1:
                    errors.append('sama próf skráð {} sinnum fyrir {}'.format(
                        groupsurveys.all().count(), studentgroup.name))
                for groupsurvey in groupsurveys.all():
                    for student in studentgroup.students.all():
                        surveyresults = SurveyResult.objects.filter(survey=groupsurvey, student=student)
                        if surveyresults.all().count() > 1:
                            errors.append('{} niðurstöður í sama prófi skráðar fyrir nemanda {} í bekk {}'.format(
                                surveyresults.all().count(),
                                student.ssn,
                                studentgroup.name,
                            ))
                        for surveyresult in surveyresults.all():
                            r = literal_eval(surveyresult.results)  # get student results
                            try:
                                survey_student_result = common_util.calc_survey_results(
                                    survey_identifier=identifier,
                                    click_values=literal_eval(r['click_values']),
                                    input_values=r['input_values'],
                                    student=student,
                                    survey_type=survey_type,
                                    transformation=transformation,
                                )
                                if not survey_student_result[0] == '':
                                    this_year_result['students_who_took_test'] += 1
                                    if int(survey_student_result[0]) >= ref_values[year][2]:
                                        this_year_result['students_over_25pct'] += 1
                                    if int(survey_student_result[0]) >= ref_values[year][1]:
                                        this_year_result['students_over_50pct'] += 1
                                    if int(survey_student_result[0]) >= ref_values[year][0]:
                                        this_year_result['students_over_90pct'] += 1
                            except:
                                pass

            ws['B' + str(index)] = this_year_result['students']
            ws['C' + str(index)] = this_year_result['students_who_took_test']
            if this_year_result['students'] > 0 and this_year_result['students_who_took_test'] > 0:
                ws['D' + str(index)] = round(
                    (this_year_result['students_who_took_test'] / this_year_result['students']) * 100
                )
                pct_over_90pct = round(
                    (this_year_result['students_over_90pct'] / this_year_result['students_who_took_test']) * 100
                )
                ws['E' + str(index)] = pct_over_90pct

                pct_over_50pct = round(
                    (this_year_result['students_over_50pct'] / this_year_result['students_who_took_test']) * 100
                )
                ws['F' + str(index)] = pct_over_50pct

                pct_over_25pct = round(
                    (this_year_result['students_over_25pct'] / this_year_result['students_who_took_test']) * 100
                )
                ws['G' + str(index)] = pct_over_25pct

            else:
                ws['D' + str(index)] = 0
                ws['E' + str(index)] = 0
                ws['F' + str(index)] = 0
                ws['G' + str(index)] = 0

            index += 1
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
        for col, value in dims.items():
            ws.column_dimensions[col].width = int(value) + 2

        chart = AreaChart()
        chart.title = "Lesfimi í {} - {}".format(title, school.name)
        chart.style = 10
        chart.width = 40
        chart.height = 20

        chart.layout = Layout(
            ManualLayout(
                xMode="edge",
                yMode="edge",
            )
        )

        chart.x_axis.title = 'Árgangur'
        chart.y_axis.title = 'Prósent'

        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 100

        cats = Reference(ws, min_col=1, min_row=2, max_row=index - 1)
        data = Reference(ws, min_col=5, min_row=1, max_col=7, max_row=index)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        bchart = BarChart()
        bchart.title = "Hlutfall nemenda sem þreyttu próf"
        bchart.style = 10
        bchart.width = 20
        bchart.height = 10

        bchart.x_axis.title = 'Árgangur'
        bchart.y_axis.title = 'Prósent'

        bchart.y_axis.scaling.min = 0
        bchart.y_axis.scaling.max = 100

        bdata = Reference(ws, min_col=4, max_col=4, min_row=2, max_row=index)
        bchart.add_data(bdata)
        bchart.legend = None
        bchart.set_categories(cats)
        ws.add_chart(bchart, "I1")

        if errors:
            index += 1
            for error in errors:
                ws['A' + str(index)] = "ATH: " + error
                ws['A' + str(index)].fill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')
                ws.merge_cells('A' + str(index) + ':F' + str(index))
                index += 1
        if index > 20:
            ws.add_chart(chart, "A" + str(index + 2))
        else:
            ws.add_chart(chart, "A22")

    # wb.save(filename='/tmp/test.xlsx')
    wb.save(response)

    return response
