from django.shortcuts import render, redirect
from django.views.generic import ListView, CreateView, DetailView, DeleteView
from django.core.urlresolvers import reverse_lazy
from django.http import HttpResponse, HttpResponseRedirect
from celery.task.control import inspect

import xlrd
import openpyxl
import collections
from itertools import chain
from django.db.models import Count, Value, CharField, Q

from kennitala import Kennitala

import common.models as cm_models
import common.mixins as cm_mixins

import supportandexception.models as sae_models
import samraemd.models as s_models
import samraemd.util as s_util
import samraemd.forms as forms
from samraemd.tasks import save_samraemd_result


def _excel_result_pre_2017(wb, school, group, year, studentgroup_id=None):
    ws = wb.get_active_sheet()
    ws.title = "Íslenska"

    ws['A1'] = 'Nemandi'
    ws['B1'] = 'Kennitala'
    ws['C1'] = 'Samræmd-Lestur'
    ws['D1'] = 'Samræmd-Málnotkun'
    ws['E1'] = 'Samræmd-Ritun'
    ws['F1'] = 'Samræmd-Heild'
    ws['G1'] = 'Raðeinkunn-Lestur'
    ws['H1'] = 'Raðeinkunn-Málnotkun'
    ws['I1'] = 'Raðeinkunn-Ritun'
    ws['J1'] = 'Raðeinkunn-Heild'
    ws['K1'] = 'Grunnskólaeinkunn-Lestur'
    ws['L1'] = 'Grunnskólaeinkunn-Málnotkun'
    ws['M1'] = 'Grunnskólaeinkunn-Ritun'
    ws['N1'] = 'Grunnskólaeinkunn-Heild'
    ws['O1'] = 'Framfaraflokkur'
    ws['P1'] = 'Framfaratexti'
    # prepare data
    studentgroup = None
    if studentgroup_id:
        studentgroup = cm_models.StudentGroup.objects.get(pk=studentgroup_id)

    results = []
    if studentgroup:
        results = s_models.SamraemdISLResult.objects.filter(
            student__in=studentgroup.students.all(),
            student_year=group,
            exam_date__year=year,
        )
    else:
        results = s_models.SamraemdISLResult.objects.filter(
            student__in=cm_models.Student.objects.filter(school=school),
            student_year=group,
            exam_date__year=year
        )

    index = 2
    for result in results:
        ws.cell('A' + str(index)).value = result.student.name
        ws.cell('B' + str(index)).value = result.student.ssn
        ws.cell('C' + str(index)).value = result.le_se.replace('.', ',')
        ws.cell('D' + str(index)).value = result.mn_se.replace('.', ',')
        ws.cell('E' + str(index)).value = result.ri_se.replace('.', ',')
        ws.cell('F' + str(index)).value = result.se.replace('.', ',')
        ws.cell('G' + str(index)).value = result.le_re.replace('.', ',')
        ws.cell('H' + str(index)).value = result.mn_re.replace('.', ',')
        ws.cell('I' + str(index)).value = result.ri_re.replace('.', ',')
        ws.cell('J' + str(index)).value = result.re.replace('.', ',')
        ws.cell('K' + str(index)).value = result.le_sg.replace('.', ',')
        ws.cell('L' + str(index)).value = result.mn_sg.replace('.', ',')
        ws.cell('M' + str(index)).value = result.ri_sg
        ws.cell('N' + str(index)).value = result.sg
        ws.cell('O' + str(index)).value = result.fm_fl.replace('.0', '')
        ws.cell('P' + str(index)).value = result.fm_txt
        index += 1
    ws = wb.create_sheet(title='Stærðfræði')

    ws['A1'] = 'Nemandi'
    ws['B1'] = 'Kennitala'
    ws['C1'] = 'Samræmd-Reikningur og aðgerðir'
    ws['D1'] = 'Samræmd-Rúmfræði og mælingar'
    ws['E1'] = 'Samræmd-Tölur og talnaskilningur'
    ws['F1'] = 'Samræmd-Heild'
    ws['G1'] = 'Raðeinkunn-Reikningur og aðgerðir'
    ws['H1'] = 'Raðeinkunn-Rúmfræði og mælingar'
    ws['I1'] = 'Raðeinkunn-Tölur og talnaskilningur'
    ws['J1'] = 'Raðeinkunn-Heild'
    ws['K1'] = 'Grunnskólaeinkunn-Reikningur og aðgerðir'
    ws['L1'] = 'Grunnskólaeinkunn-Rúmfræði og mælingar'
    ws['M1'] = 'Grunnskólaeinkunn-Tölur og talnaskilningur'
    ws['N1'] = 'Grunnskólaeinkunn-Heild'
    ws['O1'] = 'Framfaraflokkur'
    ws['P1'] = 'Framfaratexti'
    ws['Q1'] = 'Orðadæmi og talnadæmi'
    # prepare data
    if studentgroup:
        results = s_models.SamraemdMathResult.objects.filter(
            student__in=studentgroup.students.all(),
            student_year=group,
            exam_date__year=year,
        )
    else:
        results = s_models.SamraemdMathResult.objects.filter(
            student__in=cm_models.Student.objects.filter(school=school),
            student_year=group,
            exam_date__year=year
        )

    index = 2
    for result in results:
        ws.cell('A' + str(index)).value = result.student.name
        ws.cell('B' + str(index)).value = result.student.ssn
        ws.cell('C' + str(index)).value = result.ra_se.replace('.', ',')
        ws.cell('D' + str(index)).value = result.rm_se.replace('.', ',')
        ws.cell('E' + str(index)).value = result.tt_se.replace('.', ',')
        ws.cell('F' + str(index)).value = result.se
        ws.cell('G' + str(index)).value = result.ra_re
        ws.cell('H' + str(index)).value = result.rm_re
        ws.cell('I' + str(index)).value = result.tt_re
        ws.cell('J' + str(index)).value = result.re
        ws.cell('K' + str(index)).value = result.ra_sg
        ws.cell('L' + str(index)).value = result.rm_sg
        ws.cell('M' + str(index)).value = result.tt_sg
        ws.cell('N' + str(index)).value = result.sg
        ws.cell('O' + str(index)).value = result.fm_fl.replace('.0', '')
        ws.cell('P' + str(index)).value = result.fm_txt
        ws.cell('Q' + str(index)).value = result.ord_talna_txt
        index += 1
    return wb


def _excel_result_post_2017_yngri(wb, school, group, year, studentgroup_id=None):
    ws = wb.get_active_sheet()
    ws.title = "Íslenska"

    ws['A1'] = 'Nemandi'
    ws['B1'] = 'Kennitala'
    ws['C1'] = 'Hæfnieinkunn'
    ws['D1'] = 'Raðeinkunn-Lestur'
    ws['E1'] = 'Raðeinkunn-Málnotkun'
    ws['F1'] = 'Raðeinkunn-Ritun'
    ws['G1'] = 'Raðeinkunn-Heild'
    ws['H1'] = 'Grunnskólaeinkunn-Lestur'
    ws['I1'] = 'Grunnskólaeinkunn-Málnotkun'
    ws['J1'] = 'Grunnskólaeinkunn-Ritun'
    ws['K1'] = 'Grunnskólaeinkunn-Heild'
    ws['L1'] = 'Framfaraflokkur'
    ws['M1'] = 'Framfaratexti'

    # prepare data
    studentgroup = None
    if studentgroup_id:
        studentgroup = cm_models.StudentGroup.objects.get(pk=studentgroup_id)

    results = []
    if studentgroup:
        results = s_models.SamraemdISLResult.objects.filter(
            student__in=studentgroup.students.all(),
            student_year=group,
            exam_date__year=year,
        )
    else:
        results = s_models.SamraemdISLResult.objects.filter(
            student__in=cm_models.Student.objects.filter(school=school),
            student_year=group,
            exam_date__year=year
        )

    index = 2
    for result in results:
        ws.cell('A' + str(index)).value = result.student.name
        ws.cell('B' + str(index)).value = result.student.ssn
        ws.cell('C' + str(index)).value = result.he
        ws.cell('D' + str(index)).value = result.le_re.replace('.', ',')
        ws.cell('E' + str(index)).value = result.mn_re.replace('.', ',')
        ws.cell('F' + str(index)).value = result.ri_re.replace('.', ',')
        ws.cell('G' + str(index)).value = result.re.replace('.', ',')
        ws.cell('H' + str(index)).value = result.le_sg.replace('.', ',')
        ws.cell('I' + str(index)).value = result.mn_sg.replace('.', ',')
        ws.cell('J' + str(index)).value = result.ri_sg
        ws.cell('K' + str(index)).value = result.sg
        ws.cell('L' + str(index)).value = result.fm_fl.replace('.0', '')
        ws.cell('M' + str(index)).value = result.fm_txt

        index += 1

    ws = wb.create_sheet(title='Stærðfræði')

    ws['A1'] = 'Nemandi'
    ws['B1'] = 'Kennitala'
    ws['C1'] = 'Hæfnieinkunn'
    ws['D1'] = 'Raðeinkunn-Reikningur og aðgerðir'
    ws['E1'] = 'Raðeinkunn-Rúmfræði og mælingar'
    ws['F1'] = 'Raðeinkunn-Tölur og talnaskilningur'
    ws['G1'] = 'Raðeinkunn-Heild'
    ws['H1'] = 'Grunnskólaeinkunn-Reikningur og aðgerðir'
    ws['I1'] = 'Grunnskólaeinkunn-Rúmfræði og mælingar'
    ws['J1'] = 'Grunnskólaeinkunn-Tölur og talnaskilningur'
    ws['K1'] = 'Grunnskólaeinkunn-Heild'
    ws['L1'] = 'Framfaraflokkur'
    ws['M1'] = 'Framfaratexti'
    ws['N1'] = 'Orðadæmi og talnadæmi'

    # prepare data
    results = []
    if studentgroup:
        results = s_models.SamraemdMathResult.objects.filter(
            student__in=studentgroup.students.all(),
            student_year=group,
            exam_date__year=year,
        )
    else:
        results = s_models.SamraemdMathResult.objects.filter(
            student__in=cm_models.Student.objects.filter(school=school),
            student_year=group,
            exam_date__year=year
        )

    index = 2
    for result in results:
        ws.cell('A' + str(index)).value = result.student.name
        ws.cell('B' + str(index)).value = result.student.ssn
        ws.cell('C' + str(index)).value = result.he
        ws.cell('D' + str(index)).value = result.ra_re
        ws.cell('E' + str(index)).value = result.rm_re
        ws.cell('F' + str(index)).value = result.tt_re
        ws.cell('G' + str(index)).value = result.re
        ws.cell('H' + str(index)).value = result.ra_sg
        ws.cell('I' + str(index)).value = result.rm_sg
        ws.cell('J' + str(index)).value = result.tt_sg
        ws.cell('K' + str(index)).value = result.sg
        ws.cell('L' + str(index)).value = result.fm_fl.replace('.0', '')
        ws.cell('M' + str(index)).value = result.fm_txt
        ws.cell('N' + str(index)).value = result.ord_talna_txt

        index += 1

    return wb


def _excel_result_post_2017_eldri(wb, school, group, year, studentgroup_id=None):
    ws = wb.get_active_sheet()
    ws.title = "Íslenska"

    ws['A1'] = 'Nemandi'
    ws['B1'] = 'Kennitala'
    ws['C1'] = 'Hæfnieinkunn'
    ws['D1'] = 'Raðeinkunn-Lestur'
    ws['E1'] = 'Raðeinkunn-Málnotkun'
    ws['F1'] = 'Raðeinkunn-Heild'
    ws['G1'] = 'Grunnskólaeinkunn-Lestur'
    ws['H1'] = 'Grunnskólaeinkunn-Málnotkun'
    ws['I1'] = 'Grunnskólaeinkunn-Heild'
    ws['J1'] = 'Framfaraflokkur'
    ws['K1'] = 'Framfaratexti'
    # prepare data
    studentgroup = None
    if studentgroup_id:
        studentgroup = cm_models.StudentGroup.objects.get(pk=studentgroup_id)

    results = []
    if studentgroup:
        results = s_models.SamraemdISLResult.objects.filter(
            student__in=studentgroup.students.all(),
            student_year=group,
            exam_date__year=year,
        )
    else:
        results = s_models.SamraemdISLResult.objects.filter(
            student__in=cm_models.Student.objects.filter(school=school),
            student_year=group,
            exam_date__year=year
        )

    index = 2
    for result in results:
        ws.cell('A' + str(index)).value = result.student.name
        ws.cell('B' + str(index)).value = result.student.ssn
        ws.cell('C' + str(index)).value = result.he
        ws.cell('D' + str(index)).value = result.le_re.replace('.', ',')
        ws.cell('E' + str(index)).value = result.mn_re.replace('.', ',')
        ws.cell('F' + str(index)).value = result.re.replace('.', ',')
        ws.cell('G' + str(index)).value = result.le_sg.replace('.', ',')
        ws.cell('H' + str(index)).value = result.mn_sg.replace('.', ',')
        ws.cell('I' + str(index)).value = result.sg
        ws.cell('J' + str(index)).value = result.fm_fl.replace('.0', '')
        ws.cell('K' + str(index)).value = result.fm_txt
        index += 1

    ws = wb.create_sheet(title='Stærðfræði')

    ws['A1'] = 'Nemandi'
    ws['B1'] = 'Kennitala'
    ws['C1'] = 'Hæfnieinkunn'
    ws['D1'] = 'Raðeinkunn-Reikningur og aðgerðir'
    ws['E1'] = 'Raðeinkunn-Rúmfræði og mælingar'
    ws['F1'] = 'Raðeinkunn-Algebra'
    ws['G1'] = 'Raðeinkunn-Hlutföll og prósentur'
    ws['H1'] = 'Raðeinkunn-Heild'
    ws['I1'] = 'Grunnskólaeinkunn-Reikningur og aðgerðir'
    ws['J1'] = 'Grunnskólaeinkunn-Rúmfræði og mælingar'
    ws['K1'] = 'Grunnskólaeinkunn-Algebra'
    ws['L1'] = 'Grunnskólaeinkunn-Hlutföll og prósentur'
    ws['M1'] = 'Grunnskólaeinkunn-Heild'
    ws['N1'] = 'Framfaraflokkur'
    ws['O1'] = 'Framfaratexti'
    ws['P1'] = 'Orðadæmi og talnadæmi'
    # prepare data
    results = []
    if studentgroup:
        results = s_models.SamraemdMathResult.objects.filter(
            student__in=studentgroup.students.all(),
            student_year=group,
            exam_date__year=year,
        )
    else:
        results = s_models.SamraemdMathResult.objects.filter(
            student__in=cm_models.Student.objects.filter(school=school),
            student_year=group,
            exam_date__year=year
        )

    index = 2
    for result in results:
        ws.cell('A' + str(index)).value = result.student.name
        ws.cell('B' + str(index)).value = result.student.ssn
        ws.cell('C' + str(index)).value = result.he
        ws.cell('D' + str(index)).value = result.ra_re
        ws.cell('E' + str(index)).value = result.rm_re
        ws.cell('F' + str(index)).value = result.al_re
        ws.cell('G' + str(index)).value = result.hp_re
        ws.cell('H' + str(index)).value = result.re
        ws.cell('I' + str(index)).value = result.ra_sg
        ws.cell('J' + str(index)).value = result.rm_sg
        ws.cell('K' + str(index)).value = result.al_sg
        ws.cell('L' + str(index)).value = result.hp_sg
        ws.cell('M' + str(index)).value = result.sg
        ws.cell('N' + str(index)).value = result.fm_fl.replace('.0', '')
        ws.cell('O' + str(index)).value = result.fm_txt
        ws.cell('P' + str(index)).value = result.ord_talna_txt

        index += 1

    ws = wb.create_sheet(title='Enska')

    ws['A1'] = 'Nemandi'
    ws['B1'] = 'Kennitala'
    ws['C1'] = 'Hæfnieinkunn'
    ws['D1'] = 'Raðeinkunn-Lestur'
    ws['E1'] = 'Raðeinkunn-Málnotkun'
    ws['F1'] = 'Raðeinkunn-Heild'
    ws['G1'] = 'Grunnskólaeinkunn-Lestur'
    ws['H1'] = 'Grunnskólaeinkunn-Málnotkun'
    ws['I1'] = 'Grunnskólaeinkunn-Heild'
    ws['J1'] = 'Framfaraflokkur'
    ws['K1'] = 'Framfaratexti'
    # prepare data
    results = []
    if studentgroup:
        results = s_models.SamraemdENSResult.objects.filter(
            student__in=studentgroup.students.all(),
            student_year=group,
            exam_date__year=year,
        )
    else:
        results = s_models.SamraemdENSResult.objects.filter(
            student__in=cm_models.Student.objects.filter(school=school),
            student_year=group,
            exam_date__year=year
        )

    index = 2
    for result in results:
        ws.cell('A' + str(index)).value = result.student.name
        ws.cell('B' + str(index)).value = result.student.ssn
        ws.cell('C' + str(index)).value = result.he
        ws.cell('D' + str(index)).value = result.le_re.replace('.', ',')
        ws.cell('E' + str(index)).value = result.mn_re.replace('.', ',')
        ws.cell('F' + str(index)).value = result.re.replace('.', ',')
        ws.cell('G' + str(index)).value = result.le_sg.replace('.', ',')
        ws.cell('H' + str(index)).value = result.mn_sg.replace('.', ',')
        ws.cell('I' + str(index)).value = result.sg
        ws.cell('J' + str(index)).value = result.fm_fl.replace('.0', '')
        ws.cell('K' + str(index)).value = result.fm_txt
        index += 1

    return wb


def excel_result(request, school_id, year, group, studentgroup_id=None):
    school = cm_models.School.objects.get(id=school_id)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename={0}-{1}-{2}-bekkur.xlsx'.format(
        school.name, year, group)
    wb = openpyxl.Workbook()
    if int(year) < 2017:
        wb = _excel_result_pre_2017(wb, school, group, year, studentgroup_id)
    else:  # year > 2016
        if int(group) < 9:
            wb = _excel_result_post_2017_yngri(wb, school, group, year, studentgroup_id)
        else:
            wb = _excel_result_post_2017_eldri(wb, school, group, year, studentgroup_id)
    wb.save(response)

    return response


class SamraemdResultAdminListing(cm_mixins.SuperUserMixin, ListView):
    model = s_models.SamraemdMathResult
    template_name = 'samraemd/samraemdresult_admin_list.html'

    def get_context_data(self, **kwargs):
        # xxx will be available in the template as the related objects
        context = super(SamraemdResultAdminListing, self).get_context_data(**kwargs)

        if 'cancel_import' in self.request.GET:
            print("Samraemd: Cancel import, removing session data")
            del(self.request.session['newdata'])

        isl_exams = s_models.SamraemdISLResult.objects.all().values(
            'exam_date', 'student_year', 'exam_code'
        ).distinct()

        math_exams = s_models.SamraemdMathResult.objects.all().values(
            'exam_date', 'student_year', 'exam_code'
        ).distinct()

        ens_exams = s_models.SamraemdENSResult.objects.all().values(
            'exam_date', 'student_year', 'exam_code'
        ).distinct()

        context['isl_exams'] = isl_exams
        context['math_exams'] = math_exams
        context['ens_exams'] = ens_exams

        jobs = []

        active_celery_tasks = inspect().active()

        for k in active_celery_tasks.keys():
            host_tasks = active_celery_tasks[k]
            for host_task in host_tasks:
                if host_task.get('name') == 'save_samraemd_result':
                    jobs.append(host_task.get('id'))

        context['jobs'] = jobs

        return context


class SamraemdMathResultListing(cm_mixins.SchoolEmployeeMixin, ListView):
    model = s_models.SamraemdMathResult

    def get_context_data(self, **kwargs):
        # xxx will be available in the template as the related objects
        context = super(SamraemdMathResultListing, self).get_context_data(**kwargs)
        school = cm_models.School.objects.get(pk=self.kwargs['school_id'])
        context['exams'] = s_models.SamraemdMathResult.objects.filter(
            student__in=cm_models.Student.objects.filter(school=school)
        ).values('exam_code').distinct()
        context['school_id'] = school.id
        return context


class SamraemdStudentGroupResultListing(cm_mixins.SchoolTeacherMixin, ListView):
    model = s_models.SamraemdMathResult
    template_name = 'samraemd/samraemdschoolresult_list.html'

    def get_context_data(self, **kwargs):
        # xxx will be available in the template as the related objects
        context = super(SamraemdStudentGroupResultListing, self).get_context_data(**kwargs)
        school = cm_models.School.objects.get(pk=self.kwargs['school_id'])
        studentgroup = cm_models.StudentGroup.objects.get(pk=self.kwargs['studentgroup_id'])

        m_dates = s_models.SamraemdMathResult.objects.filter(
            student__in=studentgroup.students.all()
        ).values('exam_date', 'student_year').distinct()
        m_yg = [(x['exam_date'].year, x['student_year']) for x in m_dates]
        i_dates = s_models.SamraemdISLResult.objects.filter(
            student__in=studentgroup.students.all()
        ).values('exam_date', 'student_year').distinct()
        i_yg = [(x['exam_date'].year, x['student_year']) for x in i_dates]
        e_dates = s_models.SamraemdENSResult.objects.filter(
            student__in=studentgroup.students.all()
        ).values('exam_date', 'student_year').distinct()
        e_yg = [(x['exam_date'].year, x['student_year']) for x in e_dates]
        dates = list(set().union(m_yg, i_yg, e_yg))

        s_dates = sorted(dates, key=lambda tup: (tup[0], tup[1]), reverse=True)

        results = []

        for year, group in s_dates:
            student_results = {}
            for result in list(chain(
                s_models.SamraemdISLResult.objects.filter(
                    student__in=cm_models.Student.objects.filter(school=school, studentgroup=studentgroup),
                    student_year=group,
                    exam_date__year=year,
                ).annotate(result_type=Value('ÍSL', CharField())),
                s_models.SamraemdMathResult.objects.filter(
                    student__in=cm_models.Student.objects.filter(school=school, studentgroup=studentgroup),
                    student_year=group,
                    exam_date__year=year,
                ).annotate(result_type=Value('STÆ', CharField())),
                s_models.SamraemdENSResult.objects.filter(
                    student__in=cm_models.Student.objects.filter(school=school, studentgroup=studentgroup),
                    student_year=group,
                    exam_date__year=year,
                ).annotate(result_type=Value('ENS', CharField())),
            )):
                if result.student in student_results:
                    student_results[result.student].append(result)
                else:
                    student_results[result.student] = [result]
            results.append((year, group, student_results))

        rawlinks = s_models.SamraemdResult.objects.filter(
            student__in=studentgroup.students.all()
        ).values('exam_date', 'student_year', 'exam_code', 'exam_name').distinct()
        context['school'] = school
        context['studentgroup'] = studentgroup
        context['results'] = results
        context['rawlinks'] = rawlinks
        return context


class SamraemdResultListing(cm_mixins.SchoolManagerMixin, ListView):
    model = s_models.SamraemdMathResult
    template_name = 'samraemd/samraemdschoolresult_list.html'

    def get_context_data(self, **kwargs):
        # xxx will be available in the template as the related objects
        context = super(SamraemdResultListing, self).get_context_data(**kwargs)
        school = cm_models.School.objects.get(pk=self.kwargs['school_id'])

        m_dates = s_models.SamraemdMathResult.objects.filter(
            student__in=school.students.all()
        ).values('exam_date', 'student_year').distinct()
        m_yg = [(x['exam_date'].year, x['student_year']) for x in m_dates]
        i_dates = s_models.SamraemdISLResult.objects.filter(
            student__in=school.students.all()
        ).values('exam_date', 'student_year').distinct()
        i_yg = [(x['exam_date'].year, x['student_year']) for x in i_dates]
        e_dates = s_models.SamraemdENSResult.objects.filter(
            student__in=school.students.all()
        ).values('exam_date', 'student_year').distinct()
        e_yg = [(x['exam_date'].year, x['student_year']) for x in e_dates]
        dates = list(set().union(m_yg, i_yg, e_yg))

        s_dates = sorted(dates, key=lambda tup: (tup[0], tup[1]), reverse=True)

        results = []

        for year, group in s_dates:
            student_results = {}
            for result in list(chain(
                s_models.SamraemdISLResult.objects.filter(
                    student__in=cm_models.Student.objects.filter(school=school),
                    student_year=group,
                    exam_date__year=year,
                ).annotate(result_type=Value('ÍSL', CharField())),
                s_models.SamraemdMathResult.objects.filter(
                    student__in=cm_models.Student.objects.filter(school=school),
                    student_year=group,
                    exam_date__year=year,
                ).annotate(result_type=Value('STÆ', CharField())),
                s_models.SamraemdENSResult.objects.filter(
                    student__in=cm_models.Student.objects.filter(school=school),
                    student_year=group,
                    exam_date__year=year,
                ).annotate(result_type=Value('ENS', CharField())),
            )):
                if result.student in student_results:
                    student_results[result.student].append(result)
                else:
                    student_results[result.student] = [result]
            results.append((year, group, student_results))

        rawlinks = s_models.SamraemdResult.objects.filter(
            student__in=school.students.all()
        ).values('exam_date', 'student_year', 'exam_code', 'exam_name').distinct()
        context['school'] = school
        context['results'] = results
        context['rawlinks'] = rawlinks
        return context


class SamraemdResultDetail(cm_mixins.SchoolTeacherMixin, DetailView):
    model = s_models.SamraemdMathResult

    def get_template_names(self, **kwargs):
        if 'einkunnablod' in self.request.path:
            if int(self.kwargs['year']) > 2016:
                return ['samraemd/samraemd_detail_print_singles_2017.html']
            return ['samraemd/samraemd_detail_print_singles.html']
        return ['samraemd/samraemdresult_detail.html']

    def get_object(self, **kwargs):
        return s_models.SamraemdMathResult.objects.first()

    def get_context_data(self, **kwargs):
        # xxx will be available in the template as the related objects
        context = super(SamraemdResultDetail, self).get_context_data(**kwargs)
        year = self.kwargs['year']
        context['year'] = year
        group = self.kwargs['group']
        context['group'] = group
        student_results = {}
        if 'school_id' in self.kwargs:
            school = cm_models.School.objects.get(pk=self.kwargs['school_id'])
            context['school'] = school
            context['school_id'] = self.kwargs['school_id']
            context['school_name'] = school.name
            if 'student_id' in self.kwargs:
                student = cm_models.Student.objects.get(pk=self.kwargs['student_id'])
                for result in list(chain(
                    s_models.SamraemdISLResult.objects.filter(
                        student=student,
                        student_year=group,
                        exam_date__year=year,
                    ).annotate(result_type=Value('ÍSL', CharField())),
                    s_models.SamraemdMathResult.objects.filter(
                        student=student,
                        student_year=group,
                        exam_date__year=year,
                    ).annotate(result_type=Value('STÆ', CharField())),
                    s_models.SamraemdENSResult.objects.filter(
                        student=student,
                        student_year=group,
                        exam_date__year=year,
                    ).annotate(result_type=Value('ENS', CharField())),
                )):
                    if result.student in student_results:
                        student_results[result.student].append(result)
                    else:
                        student_results[result.student] = [result]
            elif 'studentgroup_id' in self.kwargs:
                studentgroup = cm_models.StudentGroup.objects.get(pk=self.kwargs['studentgroup_id'])
                for result in list(chain(
                    s_models.SamraemdISLResult.objects.filter(
                        student__in=studentgroup.students.all(),
                        student_year=group,
                        exam_date__year=year,
                    ).annotate(result_type=Value('ÍSL', CharField())),
                    s_models.SamraemdMathResult.objects.filter(
                        student__in=studentgroup.students.all(),
                        student_year=group,
                        exam_date__year=year,
                    ).annotate(result_type=Value('STÆ', CharField())),
                    s_models.SamraemdENSResult.objects.filter(
                        student__in=studentgroup.students.all(),
                        student_year=group,
                        exam_date__year=year,
                    ).annotate(result_type=Value('ENS', CharField())),
                )):
                    if result.student in student_results:
                        student_results[result.student].append(result)
                    else:
                        student_results[result.student] = [result]
            else:
                for result in list(chain(
                    s_models.SamraemdISLResult.objects.filter(
                        student__in=cm_models.Student.objects.filter(school=school),
                        student_year=group,
                        exam_date__year=year,
                    ).annotate(result_type=Value('ÍSL', CharField())),
                    s_models.SamraemdMathResult.objects.filter(
                        student__in=cm_models.Student.objects.filter(school=school),
                        student_year=group,
                        exam_date__year=year,
                    ).annotate(result_type=Value('STÆ', CharField())),
                    s_models.SamraemdENSResult.objects.filter(
                        student__in=cm_models.Student.objects.filter(school=school),
                        student_year=group,
                        exam_date__year=year,
                    ).annotate(result_type=Value('ENS', CharField())),
                )):
                    if result.student in student_results:
                        student_results[result.student].append(result)
                    else:
                        student_results[result.student] = [result]
        else:
            if self.request.user.is_superuser:
                if 'stæ' in self.request.path:
                    for result in s_models.SamraemdMathResult.objects.filter(
                        student_year=group,
                        exam_date__year=year
                    ):
                        if result.student in student_results:
                            student_results[result.student].append(result)
                        else:
                            student_results[result.student] = [result]
                elif 'isl' in self.request.path:
                    for result in s_models.SamraemdISLResult.objects.filter(
                        student_year=group,
                        exam_date__year=year
                    ):
                        if result.student in student_results:
                            student_results[result.student].append(result)
                        else:
                            student_results[result.student] = [result]
                elif 'ens' in self.request.path:
                    for result in s_models.SamraemdENSResult.objects.filter(
                        student_year=group,
                        exam_date__year=year
                    ):
                        if result.student in student_results:
                            student_results[result.student].append(result)
                        else:
                            student_results[result.student] = [result]
        context['student_results'] = collections.OrderedDict(sorted(student_results.items(), key=lambda x: (x[0].name)))

        return context


class SamraemdResultCreate(cm_mixins.SuperUserMixin, CreateView):
    model = s_models.SamraemdMathResult
    form_class = forms.SamraemdMathResultForm
    template_name = "samraemd/form_import.html"

    def post(self, *args, **kwargs):
        if (self.request.FILES):
            u_file = self.request.FILES['file'].name
            extension = u_file.split(".")[-1]
            exam_code = self.request.POST.get('exam_code').strip()
            exam_date = self.request.POST.get('exam_date').strip()
            student_year = self.request.POST.get('student_year').strip()
            exam_type = self.request.POST.get('exam_type').strip()

            data = []
            errors = []
            if extension == 'xlsx':
                input_excel = self.request.FILES['file']
                book = xlrd.open_workbook(file_contents=input_excel.read())

                for sheetsnumber in range(book.nsheets):
                    sheet = book.sheet_by_index(sheetsnumber)
                    cols = []
                    # First line is reserved for a header with field names
                    for col in range(0, sheet.ncols):
                        # Get column names based on header
                        col_name = str(sheet.cell_value(0, col))
                        cols.append(col_name)

                        fieldlist = []
                        if exam_type == 'STÆ':
                            fieldlist = [f.name for f in s_models.SamraemdMathResult._meta.get_fields()]
                        elif exam_type == 'ÍSL':
                            fieldlist = [f.name for f in s_models.SamraemdISLResult._meta.get_fields()]
                        elif exam_type == 'ENS':
                            fieldlist = [f.name for f in s_models.SamraemdENSResult._meta.get_fields()]

                        if col_name not in fieldlist:
                            if not col_name == 'ssn':
                                errors.append({
                                    'text': 'Dálkur {} er ekki til'.format(col_name),
                                    'row': 0,
                                })

                    # Run through each line
                    for row in range(1, sheet.nrows):
                        rowerrors = []
                        results_dict = {}
                        # Run through each column of the line
                        for col in range(0, len(cols)):
                            col_type = cols[col]
                            col_value = str(sheet.cell_value(row, col)).strip()

                            if col_value:
                                if col_type == 'ssn':
                                    kt = Kennitala(col_value)
                                    if not kt.validate() or not kt.is_personal(col_value):
                                        rowerrors.append({
                                            'text': 'Kennitala ekki rétt slegin inn: {}'.format(col_value),
                                            'row': row,
                                        })

                                    if not cm_models.Student.objects.filter(ssn=col_value).exists():
                                        rowerrors.append({
                                            'text': 'Nemandi ekki til í skólagátt: {}'.format(col_value),
                                            'row': row,
                                        })
                                try:
                                    if col_type == 're' or col_type.endswith('_re'):
                                        if int(col_value) > 100 or int(col_value) < 0:
                                            rowerrors.append({
                                                'text': '{}: Utan raðeinkunnarsviðs: {}'.format(col_type, col_value),
                                                'row': row,
                                            })
                                    elif col_type == 'he':
                                        if col_value not in ['A', 'B', 'B+', 'C', 'C+', 'D']:
                                            rowerrors.append({
                                                'text': '{}: Utan Hæfnieinkunnarsviðs: {}'.format(col_type, col_value),
                                                'row': row,
                                            })
                                    elif col_type == 'se' or col_type.endswith('_se'):
                                        if int(col_value) > 10 or int(col_value) < 0:
                                            rowerrors.append({
                                                'text': '{}: Utan samræmdrareinkunnarsviðs: {}'.format(
                                                    col_type,
                                                    col_value,
                                                ),
                                                'row': row,
                                            })
                                    elif col_type == 'sg' or col_type.endswith('_sg'):
                                        col
                                        if int(col_value) > 60 or int(col_value) < 0:
                                            rowerrors.append({
                                                'text': '{}: Utan grunnskólaeinnkunnarsviðs: {}'.format(
                                                    col_type,
                                                    col_value,
                                                ),
                                                'row': row,
                                            })
                                except:
                                    rowerrors.append({
                                        'text': "Villa í gildi {}: {}".format(col_type, col_value),
                                        'row': row,
                                    })

                                results_dict[col_type] = col_value

                        results_dict['exam_code'] = exam_code
                        results_dict['exam_date'] = exam_date
                        results_dict['student_year'] = student_year
                        results_dict['exam_type'] = exam_type

                        if 'ssn' not in results_dict or not results_dict['ssn']:
                            rowerrors.append({
                                'text': 'Engin kennitala',
                                'row': row
                            })

                        if rowerrors:
                            errors += rowerrors

                        data.append(results_dict)

            self.request.session['newdata'] = data
            return render(self.request, 'excel_verify_import.html', {
                'data': data,
                'errors': errors,
                'cancel_url': reverse_lazy('samraemd:result_admin_listing') + "?cancel_import",
            })
        else:
            newdata = self.request.session['newdata']
            del(self.request.session['newdata'])
            print("Calling save_samraemd_result for import")
            save_samraemd_result.delay(newdata)
        redir = self.get_success_url()
        print("Redirecting to {}".format(redir))
        return HttpResponseRedirect(redir)

    def get_success_url(self):
        return reverse_lazy('samraemd:result_admin_listing')

    def get_context_data(self, **kwargs):
        # xxx will be available in the template as the related objects
        context = super(SamraemdResultCreate, self).get_context_data(**kwargs)
        context['type'] = 'STÆ'
        return context


class SamraemdResultDelete(cm_mixins.SchoolManagerMixin, DeleteView):
    model = s_models.SamraemdResult
    template_name = "schools/confirm_delete.html"

    def get_success_url(self):
        return reverse_lazy('schools:school_listing')

    def get_object(self):
        return s_models.SamraemdResult.objects.filter(
            exam_code=self.kwargs['exam_code'],
            student_year=self.kwargs['group'],
            exam_date__year=self.kwargs['year'])


class SamraemdMathResultDelete(cm_mixins.SchoolManagerMixin, DeleteView):
    model = s_models.SamraemdMathResult
    template_name = "schools/confirm_delete.html"

    def get_success_url(self):
        return reverse_lazy('samraemd:result_admin_listing')

    def get_object(self):
        return s_models.SamraemdMathResult.objects.filter(exam_code=self.kwargs['exam_code'])


class SamraemdISLResultListing(cm_mixins.SchoolEmployeeMixin, ListView):
    model = s_models.SamraemdISLResult

    def get_context_data(self, **kwargs):
        # xxx will be available in the template as the related objects
        context = super(SamraemdISLResultListing, self).get_context_data(**kwargs)
        school = cm_models.School.objects.get(pk=self.kwargs['school_id'])
        context['exams'] = s_models.SamraemdISLResult.objects.filter(
            student__in=cm_models.Student.objects.filter(school=school)
        ).values('exam_code').distinct()
        context['school_id'] = school.id
        return context


class SamraemdISLResultDelete(cm_mixins.SchoolManagerMixin, DeleteView):
    model = s_models.SamraemdISLResult
    template_name = "schools/confirm_delete.html"

    def get_success_url(self):
        return reverse_lazy('samraemd:result_admin_listing')

    def get_object(self):
        return s_models.SamraemdISLResult.objects.filter(exam_code=self.kwargs['exam_code'])


class SamraemdENSResultListing(cm_mixins.SchoolEmployeeMixin, ListView):
    model = s_models.SamraemdENSResult

    def get_context_data(self, **kwargs):
        # xxx will be available in the template as the related objects
        context = super(SamraemdENSResultListing, self).get_context_data(**kwargs)
        school = cm_models.School.objects.get(pk=self.kwargs['school_id'])
        context['exams'] = s_models.SamraemdENSResult.objects.filter(
            student__in=cm_models.Student.objects.filter(school=school)
        ).values('exam_code').distinct()
        context['school_id'] = school.id
        return context


class SamraemdENSResultDelete(cm_mixins.SchoolManagerMixin, DeleteView):
    model = s_models.SamraemdENSResult
    template_name = "schools/confirm_delete.html"

    def get_success_url(self):
        return reverse_lazy('samraemd:result_admin_listing')

    def get_object(self):
        return s_models.SamraemdENSResult.objects.filter(exam_code=self.kwargs['exam_code'])


def excel_for_principals(request, school_id):
    school = cm_models.School.objects.get(pk=school_id)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Samræmd próf 9 og 10 bekk {} 2017.xlsx'.format(school.name)

    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Samræmt próf 2017"

    students_9 = cm_models.Student.objects.filter(
        school=school,
        studentgroup__student_year='9',
    )
    students_10 = cm_models.Student.objects.filter(
        school=school,
        studentgroup__student_year='10',
    )

    results_math_9 = s_models.SamraemdMathResult.objects.filter(
        student__in=students_9.all(),
        exam_code="9_bekkur_stærðfræði_2017",
    )
    results_math_10 = s_models.SamraemdMathResult.objects.filter(
        student__in=students_10.all(),
        exam_code="10_bekkur_stærðfræði_2017",
    )

    results_isl_9 = s_models.SamraemdISLResult.objects.filter(
        student__in=students_9.all(),
        exam_code="9_bekkur_íslenska_2017",
    )
    results_isl_10 = s_models.SamraemdISLResult.objects.filter(
        student__in=students_10.all(),
        exam_code="10_bekkur_íslenska_2017",
    )

    results_ens_9 = s_models.SamraemdENSResult.objects.filter(
        student__in=students_9.all(),
        exam_code="9_bekkur_enska_2017",
    )
    results_ens_10 = s_models.SamraemdENSResult.objects.filter(
        student__in=students_10.all(),
        exam_code="10_bekkur_enska_2017",
    )

    num_math_exc_9 = sae_models.Exceptions.objects.filter(student__in=students_9.all(), exam__contains='3').count()
    num_isl_exc_9 = sae_models.Exceptions.objects.filter(student__in=students_9.all(), exam__contains='1').count()
    num_ens_exc_9 = sae_models.Exceptions.objects.filter(student__in=students_9.all(), exam__contains='2').count()

    num_math_exc_10 = sae_models.Exceptions.objects.filter(student__in=students_10.all(), exam__contains='3').count()
    num_isl_exc_10 = sae_models.Exceptions.objects.filter(student__in=students_10.all(), exam__contains='1').count()
    num_ens_exc_10 = sae_models.Exceptions.objects.filter(student__in=students_10.all(), exam__contains='2').count()

    num_math_supp_9 = sae_models.SupportResource.objects.filter(student__in=students_9.all()).filter(
        Q(reading_assistance__contains='3') | Q(interpretation__contains='3') | Q(longer_time__contains='3')
    ).count()
    num_isl_supp_9 = sae_models.SupportResource.objects.filter(student__in=students_9.all()).filter(
        Q(reading_assistance__contains='1') | Q(interpretation__contains='1') | Q(longer_time__contains='1')
    ).count()
    num_ens_supp_9 = sae_models.SupportResource.objects.filter(student__in=students_9.all()).filter(
        Q(reading_assistance__contains='2') | Q(interpretation__contains='2') | Q(longer_time__contains='2')
    ).count()

    num_math_supp_10 = sae_models.SupportResource.objects.filter(student__in=students_10.all()).filter(
        Q(reading_assistance__contains='3') | Q(interpretation__contains='3') | Q(longer_time__contains='3')
    ).count()
    num_isl_supp_10 = sae_models.SupportResource.objects.filter(student__in=students_10.all()).filter(
        Q(reading_assistance__contains='1') | Q(interpretation__contains='1') | Q(longer_time__contains='1')
    ).count()
    num_ens_supp_10 = sae_models.SupportResource.objects.filter(student__in=students_10.all()).filter(
        Q(reading_assistance__contains='2') | Q(interpretation__contains='2') | Q(longer_time__contains='2')
    ).count()

    index = 1
    ws['A' + str(index)] = "Fjöldi nemenda í 9. bekk"
    ws['B' + str(index)] = students_9.count()
    index += 1
    ws['A' + str(index)] = "Fjöldi nemenda í 10. bekk"
    ws['B' + str(index)] = students_10.count()
    index += 2

    ws['A' + str(index)] = "Fjöldi nemenda sem þreytti próf"
    ws['B' + str(index)] = "Stærðfræði"
    ws['C' + str(index)] = "Íslenska"
    ws['D' + str(index)] = "Enska"
    index += 1
    ws['A' + str(index)] = "9. bekkur"
    ws['B' + str(index)] = results_math_9.count()
    ws['C' + str(index)] = results_isl_9.count()
    ws['D' + str(index)] = results_ens_9.count()
    index += 1
    ws['A' + str(index)] = "10. bekkur"
    ws['B' + str(index)] = results_math_10.count()
    ws['C' + str(index)] = results_isl_10.count()
    ws['D' + str(index)] = results_ens_10.count()
    index += 2

    ws['A' + str(index)] = "Fjöldi nemenda með undanþágu"
    ws['B' + str(index)] = "Stærðfræði"
    ws['C' + str(index)] = "Íslenska"
    ws['D' + str(index)] = "Enska"
    index += 1
    ws['A' + str(index)] = "9. bekkur"
    ws['B' + str(index)] = num_math_exc_9
    ws['C' + str(index)] = num_isl_exc_9
    ws['D' + str(index)] = num_ens_exc_9
    index += 1
    ws['A' + str(index)] = "10. bekkur"
    ws['B' + str(index)] = num_math_exc_10
    ws['C' + str(index)] = num_isl_exc_10
    ws['D' + str(index)] = num_ens_exc_10
    index += 2

    ws['A' + str(index)] = "Fjöldi nemenda með stuðning"
    ws['B' + str(index)] = "Stærðfræði"
    ws['C' + str(index)] = "Íslenska"
    ws['D' + str(index)] = "Enska"
    index += 1
    ws['A' + str(index)] = "9. bekkur"
    ws['B' + str(index)] = num_math_supp_9
    ws['C' + str(index)] = num_isl_supp_9
    ws['D' + str(index)] = num_ens_supp_9
    index += 1
    ws['A' + str(index)] = "10. bekkur"
    ws['B' + str(index)] = num_math_supp_10
    ws['C' + str(index)] = num_isl_supp_10
    ws['D' + str(index)] = num_ens_supp_10
    index += 2

    for grade in ["A", "B+", "B", "C+", "C", "D"]:
        ws['A' + str(index)] = "Fjöldi nemenda með {}".format(grade)
        ws['B' + str(index)] = "Stærðfræði"
        ws['C' + str(index)] = "Íslenska"
        ws['D' + str(index)] = "Enska"
        index += 1
        ws['A' + str(index)] = "9. bekkur"
        ws['B' + str(index)] = results_math_9.filter(he=grade).count()
        ws['C' + str(index)] = results_isl_9.filter(he=grade).count()
        ws['D' + str(index)] = results_ens_9.filter(he=grade).count()
        index += 1
        ws['A' + str(index)] = "10. bekkur"
        ws['B' + str(index)] = results_math_10.filter(he=grade).count()
        ws['C' + str(index)] = results_isl_10.filter(he=grade).count()
        ws['D' + str(index)] = results_ens_10.filter(he=grade).count()
        index += 2

    sg_einkunnir_math_9 = [int(x) for x in results_math_9.values_list('sg', flat=True)]
    sg_einkunnir_math_10 = [int(x) for x in results_math_10.values_list('sg', flat=True)]
    sg_einkunnir_isl_9 = [int(x) for x in results_isl_9.values_list('sg', flat=True)]
    sg_einkunnir_isl_10 = [int(x) for x in results_isl_10.values_list('sg', flat=True)]
    sg_einkunnir_ens_9 = [int(x) for x in results_ens_9.values_list('sg', flat=True)]
    sg_einkunnir_ens_10 = [int(x) for x in results_ens_10.values_list('sg', flat=True)]

    ws['A' + str(index)] = "Meðal grunnskólaeinkunn"
    ws['B' + str(index)] = "Stærðfræði"
    ws['C' + str(index)] = "Íslenska"
    ws['D' + str(index)] = "Enska"
    index += 1
    ws['A' + str(index)] = "9. bekkur"
    ws['B' + str(index)] = round(sum(sg_einkunnir_math_9) / len(sg_einkunnir_math_9))
    ws['C' + str(index)] = round(sum(sg_einkunnir_isl_9) / len(sg_einkunnir_isl_9))
    ws['D' + str(index)] = round(sum(sg_einkunnir_ens_9) / len(sg_einkunnir_ens_9))
    index += 1
    ws['A' + str(index)] = "10. bekkur"
    ws['B' + str(index)] = round(sum(sg_einkunnir_math_10) / len(sg_einkunnir_math_10))
    ws['C' + str(index)] = round(sum(sg_einkunnir_isl_10) / len(sg_einkunnir_isl_10))
    ws['D' + str(index)] = round(sum(sg_einkunnir_ens_10) / len(sg_einkunnir_ens_10))

    # B. SG Einkunn /"Normaldreifð" Meðaltal skólans og landsmeðaltal

    # Fix column widths
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = int(value) + 2

    wb.save(response)
    return response


class RawDataCreate(cm_mixins.SuperUserMixin, CreateView):
    model = s_models.SamraemdResult
    form_class = forms.SamraemdISLResultForm
    template_name = "samraemd/form_import_raw.html"

    def post(self, *args, **kwargs):
        row_offset = 5
        if(self.request.FILES):
            exam_code = self.request.POST.get('exam_code').strip()
            exam_name = self.request.POST.get('exam_name').strip()
            exam_date = self.request.POST.get('exam_date').strip()
            student_year = self.request.POST.get('student_year').strip()
            result_length = self.request.POST.get('result_length').strip()

            try:
                input_excel = self.request.FILES['file']
                book = xlrd.open_workbook(file_contents=input_excel.read())
                for sheetsnumber in range(book.nsheets):
                    sheet = book.sheet_by_index(sheetsnumber)
                    for row in range(row_offset, sheet.nrows):
                        student = cm_models.Student.objects.filter(
                            ssn=str(sheet.cell_value(row, 0)).strip())  # student already exists
                        if student:
                            # check if results for student exists, create if not, otherwise update
                            result_data = s_util.get_results_from_sheet(sheet, row, result_length)
                            results = s_models.SamraemdResult.objects.filter(
                                student=student, exam_code=exam_code)
                            results_dict = {
                                'student': student.first(),
                                'exam_code': exam_code,
                                'exam_name': exam_name,
                                'exam_date': exam_date,
                                'student_year': student_year,
                                'result_length': result_length,
                                'result_data': result_data
                            }
                            if results:
                                results.update(**results_dict)
                            else:
                                results = s_models.SamraemdResult.objects.create(**results_dict)
                        else:
                            # student not found
                            pass  # for now
            except Exception as e:
                return render(
                    self.request,
                    'samraemd/form_import_raw.html',
                    {'error': 'Dálkur ekki til, reyndu aftur'}
                )

        return redirect(self.get_success_url())

    def get_success_url(self):
        exam_code = self.request.POST.get('exam_code').strip()
        if '1' in exam_code:
            if 'school_id' in self.kwargs:
                return reverse_lazy(
                    'samraemd:isl_listing',
                    kwargs={'school_id': self.kwargs['school_id']}
                )
            return reverse_lazy('samraemd:isl_admin_listing')
        if '3' in exam_code:
            if 'school_id' in self.kwargs:
                return reverse_lazy(
                    'samraemd:math_listing',
                    kwargs={'school_id': self.kwargs['school_id']}
                )
            return reverse_lazy('samraemd:math_admin_listing')


class SamraemdRawResultDetail(cm_mixins.SchoolManagerMixin, DetailView):
    model = s_models.SamraemdResult

    def get_template_names(self, **kwargs):
        if 'einkunnablod' in self.request.path:
            return ['samraemd/samraemd_detail_raw_print_singles.html']
        return ['samraemd/samraemdresult_detail_raw.html']

    def get_object(self, **kwargs):
        return s_models.SamraemdMathResult.objects.first()

    def get_context_data(self, **kwargs):
        # xxx will be available in the template as the related objects
        context = super(SamraemdRawResultDetail, self).get_context_data(**kwargs)
        year = self.kwargs['year']
        context['year'] = year
        group = self.kwargs['group']
        context['group'] = group
        student_results = {}
        student_group = {}
        col_names = {}
        if 'school_id' in self.kwargs:
            school = cm_models.School.objects.get(pk=self.kwargs['school_id'])
            context['school'] = school
            context['school_id'] = self.kwargs['school_id']
            context['school_name'] = school.name
            if 'student_id' in self.kwargs:
                student = cm_models.Student.objects.get(pk=self.kwargs['student_id'])
                results = s_models.SamraemdResult.objects.filter(
                    student=student).filter(
                    exam_code=self.kwargs['exam_code']).filter(
                    student_year=group).filter(
                    exam_date__year=year)
                number_of_loops = results.annotate(total=Count('exam_code'))
            else:
                number_of_loops = s_models.SamraemdResult.objects.all().values(
                    'result_length', 'exam_code', 'exam_name').filter(
                    exam_date__year=year, student_year=group).annotate(total=Count('exam_code'))
                results = s_models.SamraemdResult.objects.filter(
                    student__in=cm_models.Student.objects.filter(school=school)).filter(
                    student_year=group).filter(exam_date__year=year)
            s_util.display_raw_results(results, student_results, student_group, col_names)

        else:
            if self.request.user.is_superuser:
                exam_code = self.kwargs['exam_code']
                context['exam_code'] = exam_code
                number_of_loops = s_models.SamraemdResult.objects.values(
                    'result_length', 'exam_code', 'exam_name').filter(
                    exam_code=exam_code, exam_date__year=year).annotate(total=Count('exam_code'))
                results = s_models.SamraemdResult.objects.filter(
                    exam_code=exam_code).filter(student_year=group).filter(exam_date__year=year)
                s_util.display_raw_results(results, student_results, student_group, col_names)

        context['student_results'] = student_results
        context['student_group'] = student_group
        context['loop_times'] = number_of_loops
        context['columns'] = col_names
        return context


def admin_result_raw_excel(request, exam_code, year, group):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=hrániðurstöður-{0}.xlsx'.format(group)
    number_of_loops = s_models.SamraemdResult.objects.filter(
        exam_code=exam_code,
        student_year=group,
        exam_date__year=year)[:1].values('result_length', 'exam_name').get()
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = number_of_loops['exam_name']

    ws['A1'] = 'Skóli'
    ws['B1'] = 'Kennitölur'
    ws['C1'] = 'Nafn'
    ws['D1'] = 'Bekkur'
    # We need the first result just to get the keys for the columns
    first_result = s_models.SamraemdResult.objects.filter(
        exam_code=exam_code,
        student_year=group,
        exam_date__year=year
    ).first().result_data.items()
    for key, result in first_result:
        ws.cell(row=1, column=int(key) + 5).value = str(result['id'])

    index = 2
    for result in s_models.SamraemdResult.objects.filter(
        exam_code=exam_code,
        student_year=group,
        exam_date__year=year
    ):
        ws.cell(row=index, column=1).value = str(
            cm_models.School.objects.get(students=result.student).name)
        ws.cell(row=index, column=2).value = str(result.student.ssn)
        ws.cell(row=index, column=3).value = str(result.student)
        ws.cell(row=index, column=4).value = str(
            cm_models.StudentGroup.objects.get(students=result.student))
        for key, values in result.result_data.items():
            ws.cell(row=index, column=int(key) + 5).value = values['value']

        index += 1

    wb.save(response)

    return response


def excel_result_raw(request, school_id, year, group):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=hrániðurstöður-{0}.xlsx'.format(group)
    wb = openpyxl.Workbook()
    number_of_loops = s_models.SamraemdResult.objects.all().values(
        'result_length', 'exam_code', 'exam_name').filter(
        exam_date__year=year, student_year=group).annotate(total=Count('exam_code'))
    for loops in number_of_loops:
        ws = wb.create_sheet()
        if int(loops['exam_code']) == 1:
            ws.title = 'Íslenska'
        elif int(loops['exam_code']) == 2:
            ws.title = 'Enska'
        elif int(loops['exam_code']) == 3:
            ws.title = 'Stærðfræði'

        ws['A1'] = 'Skóli'
        ws['B1'] = 'Kennitölur'
        ws['C1'] = 'Nafn'
        ws['D1'] = 'bekkur'
        # We need the first result just to get the keys for the columns
        for result in s_models.SamraemdResult.objects.filter(
            student__in=cm_models.Student.objects.filter(school=school_id),
            exam_code=loops['exam_code'],
            student_year=group,
            exam_date__year=year
        ):
            for key, values in result.result_data.items():
                ws.cell(row=1, column=int(key) + 5).value = values['id']
        index = 2
        for result in s_models.SamraemdResult.objects.filter(
            student__in=cm_models.Student.objects.filter(school=school_id),
            exam_code=loops['exam_code'],
            student_year=group,
            exam_date__year=year
        ):
            ws.cell(row=index, column=1).value = str(
                cm_models.School.objects.get(students=result.student).name)
            ws.cell(row=index, column=2).value = str(result.student.ssn)
            ws.cell(row=index, column=3).value = str(result.student)
            try:
                ws.cell(row=index, column=4).value = str(
                    cm_models.StudentGroup.objects.get(students=result.student))
            except:
                ws.cell(row=index, column=4).value = "Enginn bekkur"
            for key, values in result.result_data.items():
                ws.cell(row=index, column=int(key) + 5).value = values['value']

            index += 1

    wb.save(response)

    return response
